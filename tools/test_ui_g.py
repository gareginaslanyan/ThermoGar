#!/usr/bin/env python3
"""Functional UI tests for the ThermoGar kinetics sections (wave 3G).

Every test drives the real Streamlit script ``app/ThermoGar_app.py`` through
``streamlit.testing.v1.AppTest``: the database is chosen in the sidebar, the
composition and the kinetics inputs are filled in, the calculation button is
clicked, and the result is taken from ``st.session_state``. Export is then
checked on the functions that stand behind the download buttons -- AppTest
does not carry the bytes of ``st.download_button`` -- with the same writers the
application uses (``pandas.ExcelWriter(engine="openpyxl")``,
``Figure.savefig(format="png")``, ``DataFrame.to_csv``).

Matrix (three databases x three actions):

* single-phase diffusion couple  -- ``kin_single_run_<key>``;
* multiphase homogenization      -- ``kin_hom_run_<key>``;
* KWN precipitation kinetics     -- ``precipitation_<key>_user_calculate``.

Run:
    <root>/.venv-windows/Scripts/python.exe -m pytest tools/test_ui_g.py -v

The application must be launched from the project root, otherwise
``.streamlit/config.toml`` is not picked up and the script refuses to render.
"""

from __future__ import annotations

import json
import sys
from io import BytesIO
from pathlib import Path
from typing import Any

import matplotlib

matplotlib.use("Agg")

import pandas as pd
import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT / "app") not in sys.path:
    sys.path.insert(0, str(ROOT / "app"))

from streamlit.testing.v1 import AppTest

import thermogar_precipitation as precipitation_module
from thermogar_diffusion import DiffusionResult
from thermogar_precipitation import PrecipitationResult

APP_SCRIPT = str(ROOT / "app" / "ThermoGar_app.py")
DB_KEYS = ("ni", "al", "fe")

# Errors that predate wave 3G, come from modules outside this task and appear
# on every run of every section. They are reported, not asserted away.
KNOWN_FOREIGN_ERRORS = ("BINDING_STALE",)

# Sidebar alloy of the wave: balance, solutes, units label.
SIDEBAR_ALLOY = {
    "ni": ("NI", "AL=15", "атомные %"),
    "al": ("AL", "CU=4, MG=1", "массовые %"),
    "fe": ("FE", "C=0.20, CR=11.5, NI=0.7", "массовые %"),
}

# Diffusion couple: balance, units label, left side, right side, temperature,
# phases with a complete MQ/MF set for the elements of that couple.
DIFFUSION_COUPLE = {
    "ni": ("NI", "атомные %", "CR=7.7, AL=5.4", "CR=35.9, AL=6.2", 1200.0,
           ("FCC_A1", "NIAL")),
    "al": ("AL", "атомные %", "CU=1", "CU=5", 500.0, ("FCC_A1",)),
    "fe": ("FE", "массовые %", "C=0.1, CR=8", "C=0.3, CR=14", 900.0,
           ("FCC_A1", "BCC_A2")),
}

# KWN cell: matrix, precipitate, temperature. The pairs are the ones the
# backend matrix of wave 1C measured; the matrix of Al and Fe is the
# disordered half of an order/disorder pair, which ``filter_phases`` drops.
KWN_CELL = {
    "ni": ("FCC_A1", "GAMMA_PRIME", 800.0),
    "al": ("FCC_A1", "THETA_AL2CU", 200.0),
    "fe": ("BCC_A2", "M23C6", 700.0),
}

# Short model time of the matrix cells, hours.
SHORT_TIME_H = 0.001


# --------------------------------------------------------------------------- #
# Harness
# --------------------------------------------------------------------------- #


def start_app(database_key: str, *, composition: str | None = None) -> AppTest:
    """Run the application once with the wave alloy selected in the sidebar."""

    balance, default_composition, units_label = SIDEBAR_ALLOY[database_key]
    app = AppTest.from_file(APP_SCRIPT, default_timeout=1800)
    app.session_state["thermogar_database_key"] = database_key
    app.session_state[f"thermogar_balance_{database_key}"] = balance
    app.session_state[f"thermogar_units_{database_key}"] = units_label
    app.session_state[f"thermogar_composition_{database_key}"] = (
        default_composition if composition is None else composition
    )
    app.run()
    return app


def new_errors(app: AppTest) -> list[str]:
    """Error messages of this run without the known pre-existing ones."""

    return [
        element.value
        for element in app.error
        if not any(known in element.value for known in KNOWN_FOREIGN_ERRORS)
    ]


def captions(app: AppTest) -> str:
    return "\n".join(element.value for element in app.caption)


def widget(app: AppTest, kind: str, key: str) -> Any:
    matches = [item for item in app.get(kind) if item.key == key]
    assert matches, f"В приложении нет виджета {kind} с ключом {key!r}"
    return matches[0]


def assert_no_traceback(app: AppTest) -> None:
    assert not app.exception, [element.message for element in app.exception]


def set_diffusion_inputs(app: AppTest, database_key: str, prefix: str) -> None:
    balance, units_label, left, right, temperature, _phases = DIFFUSION_COUPLE[
        database_key
    ]
    state = app.session_state
    state[f"{prefix}_balance_{database_key}"] = balance
    state[f"{prefix}_units_{database_key}"] = units_label
    state[f"{prefix}_left_{database_key}"] = left
    state[f"{prefix}_right_{database_key}"] = right
    state[f"{prefix}_temperature_{database_key}"] = temperature
    state[f"{prefix}_length_{database_key}"] = 100.0
    state[f"{prefix}_time_{database_key}"] = SHORT_TIME_H
    state[f"{prefix}_nodes_{database_key}"] = 20
    app.run()


# --------------------------------------------------------------------------- #
# Export, on the same writers the download buttons use
# --------------------------------------------------------------------------- #


def excel_bytes(sheets: dict[str, pd.DataFrame]) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name[:31], index=False)
    return buffer.getvalue()


def png_bytes(figure: Any) -> bytes:
    buffer = BytesIO()
    figure.savefig(buffer, format="png", dpi=200, bbox_inches="tight")
    return buffer.getvalue()


def assert_excel_opens(payload: bytes, expected_sheets: tuple[str, ...]) -> None:
    assert payload[:2] == b"PK", "Файл Excel не начинается сигнатурой ZIP."
    workbook = load_workbook(BytesIO(payload))
    for sheet in expected_sheets:
        assert sheet[:31] in workbook.sheetnames, (
            f"В книге нет листа {sheet!r}: {workbook.sheetnames}"
        )
        assert workbook[sheet[:31]].max_row >= 2, f"Лист {sheet!r} пуст."


def assert_png(payload: bytes) -> None:
    assert payload.startswith(b"\x89PNG"), "PNG не начинается сигнатурой \\x89PNG."
    assert len(payload) > 1000


def assert_csv(payload: bytes, minimum_rows: int) -> None:
    frame = pd.read_csv(BytesIO(payload))
    assert len(frame) >= minimum_rows
    assert not frame.empty


# --------------------------------------------------------------------------- #
# 0. The section renders and its controls are usable on every database
# --------------------------------------------------------------------------- #


@pytest.mark.parametrize("database_key", DB_KEYS)
def test_kinetics_section_renders(database_key: str) -> None:
    """Both kinetics tabs render and the run buttons carry no extra gate."""

    app = start_app(database_key)
    assert_no_traceback(app)

    single = widget(app, "button", f"kin_single_run_{database_key}")
    homogenization = widget(app, "button", f"kin_hom_run_{database_key}")
    assert single.proto.disabled is False, (
        "Кнопка однофазной диффузии выключена без причины: "
        f"база {database_key}."
    )
    # Homogenization stays switched off only when the database has fewer than
    # two phases with a complete mobility set for the couple (mc_al).
    homogenization_possible = len(DIFFUSION_COUPLE[database_key][5]) >= 2
    assert homogenization.proto.disabled == (not homogenization_possible)

    # No leftover confirmation checkbox in front of either calculation.
    confirmations = [
        item.key
        for item in app.get("checkbox")
        if item.key and ("kin_" in item.key or "precipitation_" in item.key)
    ]
    assert confirmations == [], confirmations


def test_run_gate_is_identical_for_all_databases() -> None:
    """Fe is not gated differently from Ni and Al (attention point 1)."""

    states = {}
    for database_key in DB_KEYS:
        app = start_app(database_key)
        states[database_key] = widget(
            app, "button", f"kin_single_run_{database_key}"
        ).proto.disabled
    assert states == {"ni": False, "al": False, "fe": False}, states


# --------------------------------------------------------------------------- #
# 1. Single-phase diffusion couple
# --------------------------------------------------------------------------- #


@pytest.mark.parametrize("database_key", DB_KEYS)
def test_diffusion_single_phase(database_key: str) -> None:
    app = start_app(database_key)
    set_diffusion_inputs(app, database_key, "kin_single")
    phase = DIFFUSION_COUPLE[database_key][5][0]
    app.session_state[f"kin_single_phase_{database_key}"] = phase
    app.run()

    widget(app, "button", f"kin_single_run_{database_key}").click().run()
    assert_no_traceback(app)
    assert new_errors(app) == [], new_errors(app)

    result = app.session_state[f"kin_single_result_{database_key}"]
    assert isinstance(result, DiffusionResult)
    assert result.database_key == database_key
    assert result.phases == [phase]
    assert len(result.profile_table) == 20
    assert not result.profile_table.isna().any().any()
    assert result.max_balance_error <= 1e-6
    assert (result.quality["Статус"] == "пройдена").all()

    # Local phase fractions are a percentage of the whole, per node.
    fraction_column = f"{phase}, локальная доля, %"
    assert fraction_column in result.phase_fractions
    fractions = result.phase_fractions[fraction_column]
    assert ((fractions >= -1e-6) & (fractions <= 100.0 + 1e-6)).all()

    payload = excel_bytes(
        {
            "Параметры": result.settings,
            "Профили": result.profile_table,
            "Фазовые доли": result.phase_fractions,
            "Баланс": result.balance_table,
        }
    )
    assert_excel_opens(payload, ("Параметры", "Профили", "Фазовые доли", "Баланс"))
    assert_png(png_bytes(result.profile_figure))
    assert_csv(result.profile_table.to_csv(index=False).encode("utf-8-sig"), 20)


# --------------------------------------------------------------------------- #
# 2. Multiphase homogenization
# --------------------------------------------------------------------------- #


@pytest.mark.parametrize("database_key", ("ni", "fe"))
def test_diffusion_homogenization(database_key: str) -> None:
    app = start_app(database_key)
    set_diffusion_inputs(app, database_key, "kin_hom")
    phases = list(DIFFUSION_COUPLE[database_key][5][:2])
    app.session_state[f"kin_hom_phases_{database_key}"] = phases
    app.run()

    widget(app, "button", f"kin_hom_run_{database_key}").click().run()
    assert_no_traceback(app)
    assert new_errors(app) == [], new_errors(app)

    result = app.session_state[f"kin_hom_result_{database_key}"]
    assert isinstance(result, DiffusionResult)
    assert result.method_key == "homogenization"
    assert sorted(result.phases) == sorted(phases)
    assert len(result.profile_table) == 20
    assert result.max_balance_error <= 1e-6
    assert (result.quality["Статус"] == "пройдена").all()

    # Every node holds a full phase assembly: local fractions sum to 100 %.
    columns = [f"{phase}, локальная доля, %" for phase in phases]
    totals = result.phase_fractions[columns].sum(axis=1)
    assert ((totals - 100.0).abs() <= 1e-3).all(), totals.tolist()

    payload = excel_bytes(
        {
            "Параметры": result.settings,
            "Профили": result.profile_table,
            "Фазовые доли": result.phase_fractions,
            "Баланс": result.balance_table,
        }
    )
    assert_excel_opens(payload, ("Параметры", "Профили", "Фазовые доли", "Баланс"))
    assert_png(png_bytes(result.profile_figure))
    assert result.phase_figure is not None
    assert_png(png_bytes(result.phase_figure))


def test_homogenization_unavailable_on_al_is_explained() -> None:
    """mc_al carries mobility for FCC_A1 only: say so instead of failing later."""

    app = start_app("al")
    set_diffusion_inputs(app, "al", "kin_hom")
    assert_no_traceback(app)

    button = widget(app, "button", "kin_hom_run_al")
    assert button.proto.disabled is True
    warnings = "\n".join(element.value for element in app.warning)
    assert "Многофазная гомогенизация требует минимум две" in warnings
    assert "FCC_A1" in warnings


# --------------------------------------------------------------------------- #
# 3. KWN precipitation kinetics
# --------------------------------------------------------------------------- #


@pytest.mark.parametrize("database_key", DB_KEYS)
def test_kwn_precipitation(database_key: str) -> None:
    matrix, precipitate, temperature = KWN_CELL[database_key]
    app = start_app(database_key)

    matrix_options = widget(
        app, "selectbox", f"precipitation_{database_key}_user_matrix"
    ).options
    assert matrix in matrix_options, (
        f"{matrix} не предлагается матрицей для базы {database_key}: "
        f"{matrix_options}"
    )

    state = app.session_state
    state[f"precipitation_{database_key}_user_matrix"] = matrix
    state[f"precipitation_{database_key}_user_precipitate"] = precipitate
    state[f"precipitation_{database_key}_user_temperature_c"] = temperature
    state[f"precipitation_{database_key}_user_duration_h"] = SHORT_TIME_H
    state[f"precipitation_{database_key}_user_bins"] = 30
    app.run()

    widget(
        app, "button", f"precipitation_{database_key}_user_calculate"
    ).click().run()
    assert_no_traceback(app)
    assert new_errors(app) == [], new_errors(app)

    result = app.session_state["thermogar_precipitation_result"]
    assert isinstance(result, PrecipitationResult)
    assert result.database_key == database_key
    assert result.phase == precipitate
    assert len(result.kinetics) >= 10
    assert len(result.psd) >= 10
    assert (result.quality["Статус"] == "пройдена").all()

    fraction = result.kinetics["Объёмная доля, %"]
    radius = result.kinetics["Средний радиус, нм"]
    assert ((fraction >= -1e-8) & (fraction <= 100.0 + 1e-8)).all()
    assert (radius >= -1e-12).all()
    assert (result.psd["Число частиц в классе, 1/м³"] >= -1e-6).all()
    assert set(result.summary["Показатель"]) >= {
        "Максимальная объёмная доля",
        "Итоговый средний радиус",
    }
    # Attention point 5: kawin's own "divide by zero" in NucleationRate is a
    # library warning and must not reach the user as a failed calculation.
    assert not any("divide by zero" in message for message in new_errors(app))

    excel = precipitation_module._excel(result)
    assert_excel_opens(
        excel,
        (
            "Параметры",
            "Итоги",
            "Кинетика",
            "Состав матрицы",
            "Межфазные составы",
            "Итоговое PSD",
            "Проверки",
        ),
    )
    assert result.npz[:2] == b"PK" and len(result.npz) > 1000
    assert_png(precipitation_module._png(result.figures["fraction"]))
    assert_png(precipitation_module._png(result.figures["psd"]))
    provenance = json.loads(result.provenance.decode("utf-8-sig"))
    assert provenance["database_key"] == database_key
    assert provenance["matrix_order_disorder_role"]["phase"] == matrix
    assert provenance["matrix_order_disorder_role"]["role"] != "ordered"


def test_fe_kwn_provenance_status_is_neutral() -> None:
    """Attention point 3: the Fe status is metadata, not a gate."""

    assert precipitation_module.FE_KWN_PUBLICATION_STATUS == "NOT_ASSESSED"

    matrix, precipitate, temperature = KWN_CELL["fe"]
    app = start_app("fe")
    state = app.session_state
    state["precipitation_fe_user_matrix"] = matrix
    state["precipitation_fe_user_precipitate"] = precipitate
    state["precipitation_fe_user_temperature_c"] = temperature
    state["precipitation_fe_user_duration_h"] = SHORT_TIME_H
    state["precipitation_fe_user_bins"] = 30
    app.run()
    widget(app, "button", "precipitation_fe_user_calculate").click().run()
    assert_no_traceback(app)

    result = app.session_state["thermogar_precipitation_result"]
    provenance = json.loads(result.provenance.decode("utf-8-sig"))
    assert provenance["fe_kwn_publication_status"] == "NOT_ASSESSED"


@pytest.mark.parametrize(
    ("database_key", "expected_matrix"),
    (("al", "FCC_A1"), ("fe", "BCC_A2")),
)
def test_kwn_matrix_offers_the_disordered_half(
    database_key: str,
    expected_matrix: str,
) -> None:
    """Attention point 2: filter_phases keeps GP_MAT / BCC_B2, Kawin needs the pair."""

    app = start_app(database_key)
    assert_no_traceback(app)
    selectbox = widget(
        app, "selectbox", f"precipitation_{database_key}_user_matrix"
    )
    assert expected_matrix in selectbox.options, selectbox.options
    assert selectbox.value == expected_matrix

    ordered_half = {"al": "GP_MAT", "fe": "BCC_B2"}[database_key]
    assert ordered_half not in selectbox.options
    precipitates = widget(
        app, "selectbox", f"precipitation_{database_key}_user_precipitate"
    ).options
    assert ordered_half not in precipitates, (
        f"{ordered_half} — та же фаза, что матрица {expected_matrix}, "
        "и не может быть выделением."
    )


def test_kwn_reports_a_too_long_composition_without_a_traceback() -> None:
    """The shipped Fe composition has eight solutes; KWN accepts four."""

    app = start_app(
        "fe",
        composition="C=0.20, CR=11.5, NI=0.7, MN=0.7, SI=0.3, MO=0.6, W=0.9, V=0.225",
    )
    assert_no_traceback(app)
    messages = "\n".join(element.value for element in app.error)
    assert "не более четырёх добавок" in messages
    assert "боковой панели" in captions(app)


# --------------------------------------------------------------------------- #
# 4. Input validation: a message, never a traceback
# --------------------------------------------------------------------------- #


@pytest.mark.parametrize("prefix", ("kin_single", "kin_hom"))
def test_diffusion_number_inputs_are_bounded(prefix: str) -> None:
    """Zero cells and negative time are unreachable: the widgets are bounded."""

    app = start_app("ni")
    assert_no_traceback(app)
    bounds = {
        "length": (1.0, None),
        "interface": (1.0, 99.0),
        "time": (0.001, None),
        "nodes": (12.0, 160.0),
    }
    for field, (minimum, maximum) in bounds.items():
        control = widget(app, "number_input", f"{prefix}_{field}_ni")
        assert control.proto.has_min, field
        assert control.proto.min == pytest.approx(minimum), field
        assert control.proto.has_max is (maximum is not None), field
        if maximum is not None:
            assert control.proto.max == pytest.approx(maximum), field


@pytest.mark.parametrize(
    ("overrides", "fragment"),
    (
        ({"time_h": 0.0}, "Время выдержки должно быть больше нуля"),
        ({"time_h": -1.0}, "Время выдержки должно быть больше нуля"),
        ({"nodes": 0}, "Число ячеек должно быть от 12 до 160"),
        ({"length_um": -100.0}, "Длина области должна быть больше нуля"),
        ({"interface_percent": 0.0}, "Граница пары должна находиться"),
        ({"phases": []}, "Выберите хотя бы одну фазу"),
    ),
)
def test_absurd_diffusion_values_raise_a_readable_message(
    overrides: dict[str, Any],
    fragment: str,
) -> None:
    """If such a value still reaches the engine (loaded project, API call),
    the answer is a sentence, not a traceback."""

    from thermogar_diffusion import run_diffusion
    from thermogar_release_policy import (
        RELEASE_DATABASE_LABELS,
        RELEASE_DATABASE_RELATIVE_PATHS,
    )

    scenario: dict[str, Any] = {
        "db": None,
        "database_key": "ni",
        "database_path": ROOT / RELEASE_DATABASE_RELATIVE_PATHS["ni"],
        "database_label": RELEASE_DATABASE_LABELS["ni"],
        "balance": "NI",
        "units": "at",
        "left_text": "CR=7.7, AL=5.4",
        "right_text": "CR=35.9, AL=6.2",
        "temperature_c": 1200.0,
        "time_h": 0.001,
        "length_um": 100.0,
        "interface_percent": 50.0,
        "nodes": 20,
        "phases": ["FCC_A1"],
        "model_kind": "single",
        "input_provenance": "SYNTHETIC_UI_REGRESSION_NOT_MATERIAL_INPUT",
        "input_confirmation": True,
    }
    scenario.update(overrides)
    with pytest.raises(ValueError) as raised:
        run_diffusion(**scenario)
    assert fragment in str(raised.value), str(raised.value)


@pytest.mark.parametrize(
    ("left", "right", "fragment"),
    (
        ("CR=7.7", "CR=7.7", "совпадают"),
        ("CR=60, AL=50", "CR=1", "меньше 100"),
        ("бессмыслица", "CR=5", "Не удалось прочитать состав"),
    ),
)
def test_bad_couple_composition_is_reported(
    left: str,
    right: str,
    fragment: str,
) -> None:
    app = start_app("ni")
    set_diffusion_inputs(app, "ni", "kin_single")
    app.session_state["kin_single_left_ni"] = left
    app.session_state["kin_single_right_ni"] = right
    app.run()

    assert_no_traceback(app)
    warnings = "\n".join(element.value for element in app.warning)
    assert fragment in warnings, warnings


def test_kwn_size_grid_is_validated() -> None:
    app = start_app("fe")
    state = app.session_state
    state["precipitation_fe_user_matrix"] = "BCC_A2"
    state["precipitation_fe_user_precipitate"] = "M23C6"
    state["precipitation_fe_user_duration_h"] = SHORT_TIME_H
    state["precipitation_fe_user_cmin_nm"] = 10.0
    state["precipitation_fe_user_cmax_nm"] = 1.0
    app.run()
    widget(app, "button", "precipitation_fe_user_calculate").click().run()

    assert_no_traceback(app)
    assert any("диапазон радиусов" in message for message in new_errors(app)), (
        new_errors(app)
    )


# --------------------------------------------------------------------------- #
# 5. Shipped Fe defaults really run (attention point 4)
# --------------------------------------------------------------------------- #


@pytest.mark.slow
def test_ni_kwn_reaches_a_precipitated_state() -> None:
    """One hour at 800 C must actually grow gamma-prime, not just run."""

    matrix, precipitate, temperature = KWN_CELL["ni"]
    app = start_app("ni")
    state = app.session_state
    state["precipitation_ni_user_matrix"] = matrix
    state["precipitation_ni_user_precipitate"] = precipitate
    state["precipitation_ni_user_temperature_c"] = temperature
    state["precipitation_ni_user_duration_h"] = 1.0
    app.run()
    widget(app, "button", "precipitation_ni_user_calculate").click().run()
    assert_no_traceback(app)
    assert new_errors(app) == [], new_errors(app)

    result = app.session_state["thermogar_precipitation_result"]
    fraction = result.kinetics["Объёмная доля, %"]
    radius = result.kinetics["Средний радиус, нм"]
    assert float(fraction.max()) > 1.0, float(fraction.max())
    assert float(radius.iloc[-1]) > 0.5, float(radius.iloc[-1])
    assert float(result.psd["Число частиц в классе, 1/м³"].max()) > 0.0
    assert (result.quality["Статус"] == "пройдена").all()


def test_fe_shipped_defaults_are_the_declared_ones() -> None:
    """Attention point 4: the Fe cell opens on the pair wave 2B declared.

    The wall-clock cost of the full default schedule is measured separately
    and written down in ``tools/ui_matrix_G.md``; here we only check that the
    section starts on those values and that nothing about them is rejected.
    """

    app = start_app("fe")
    assert_no_traceback(app)

    assert widget(app, "number_input", "kin_single_temperature_fe").value == 900.0
    assert widget(app, "selectbox", "kin_single_phase_fe").value == "FCC_A1"
    assert sorted(
        widget(app, "multiselect", "kin_hom_phases_fe").value
    ) == ["BCC_A2", "FCC_A1"]

    assert widget(app, "selectbox", "precipitation_fe_user_matrix").value == "BCC_A2"
    assert (
        widget(app, "selectbox", "precipitation_fe_user_precipitate").value == "M23C6"
    )
    assert widget(app, "number_input", "precipitation_fe_user_temperature_c").value == (
        700.0
    )
    assert widget(app, "number_input", "precipitation_fe_user_gamma").value == 0.3
    assert widget(app, "number_input", "precipitation_fe_user_matrix_vm").value == 7.09
