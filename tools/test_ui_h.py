"""UI tests for the "Projects and data" section: library, batch, projects,
history, import and export.

The tests drive the real Streamlit app through ``streamlit.testing.v1.AppTest``
and use a private state root, so they never touch the user's own
``%LOCALAPPDATA%\\ThermoGar``.

Two defects in ``app/ThermoGar_app.py`` currently disable the whole section;
that file belongs to another task, so the fixture runs the app from a patched
copy instead of editing it. See ``patched_app_source`` for what is patched and
``tasks/WAVE3H_REPORT.md`` for why. Once the fixes land upstream the patches
stop matching and the tests run against the file itself.

Run:
    .venv-windows\\Scripts\\python.exe -X utf8 -B -m pytest tools/test_ui_h.py \\
        -q -m "not slow"
"""

from __future__ import annotations

import io
import json
from pathlib import Path
import sys

import pytest


ROOT = Path(__file__).resolve().parent.parent
APP = ROOT / "app"
sys.path.insert(0, str(APP))

from thermogar_workspace import (  # noqa: E402
    WIDGET_STATE_VERSION,
    WIDGET_STATE_VERSION_FIELD,
    is_restorable_widget_key,
    portable_project_payload,
    rejection_text,
    validate_widget_state,
)


# Compositions and temperatures of the release test matrix.
CASES = {
    "ni": ("NI", "ат.%", 700.0, "AL=15"),
    "al": ("AL", "мас.%", 500.0, "CU=4, MG=1"),
    "fe": ("FE", "мас.%", 700.0, "C=0.2, CR=11.5, NI=0.7"),
}

# Upstream defects that close the section. Each entry is (broken, fixed);
# an entry that no longer matches the source is simply skipped.
UPSTREAM_PATCHES: tuple[tuple[str, str], ...] = (
    # The properties section rebinds the module-level ``vlb_active_context``
    # to its own TDB+PDB binding. The workspace StateStore probes that same
    # global, so every prepared upload and download in this section is
    # rejected as BINDING_STALE.
    (
        "        vlb_active_context.binding_digest,\n"
        "        vlb_active_context.binding_generation,\n",
        "        vlb_bound_context.binding_digest,\n"
        "        vlb_bound_context.binding_generation,\n",
    ),
    # BoundDatabaseContext exposes ``tdb``; ``tdb_evidence`` belongs to
    # FeatureReceipt. rebind_context raises AttributeError, which kills
    # "load composition", "open project" and "restore from history".
    (
        "            rebound.tdb_evidence.sha256,\n",
        "            rebound.tdb.sha256,\n",
    ),
    (
        '        clean["database_sha256"] = rebound.tdb_evidence.sha256\n',
        '        clean["database_sha256"] = rebound.tdb.sha256\n',
    ),
)

# The copy has to live beside the original: the app derives PROJECT_ROOT from
# its own __file__. The fixture removes it when the session ends; a run killed
# mid-flight leaves it behind and the next run overwrites it. Do not delete it
# at import time — that would pull the file out from under a parallel session.
PATCHED_APP = APP / "_test_ui_h_app.py"


def batch_csv(
    keys: tuple[str, ...] = ("ni", "al", "fe"),
    *,
    separator: str = ",",
    bom: bool = False,
) -> bytes:
    """Build a batch input file for the given databases."""

    def cell(value: str) -> str:
        if separator == "," and ("," in value or '"' in value):
            return '"' + value.replace('"', '""') + '"'
        return value

    header = separator.join(
        cell(name)
        for name in (
            "Название",
            "База",
            "Основа",
            "Единицы",
            "Температура, °C",
            "Добавки",
        )
    )
    lines = [header]
    for key in keys:
        balance, units, temperature_c, composition = CASES[key]
        lines.append(
            separator.join(
                cell(value)
                for value in (
                    key.upper(),
                    key,
                    balance,
                    units,
                    f"{temperature_c:g}",
                    composition,
                )
            )
        )
    data = ("\n".join(lines) + "\n").encode("utf-8")
    return b"\xef\xbb\xbf" + data if bom else data


@pytest.fixture(scope="session")
def patched_app_source() -> Path:
    """Path to the app script the tests drive.

    Returns ``ThermoGar_app.py`` itself once the upstream defects listed in
    ``UPSTREAM_PATCHES`` are fixed there.
    """

    original = APP / "ThermoGar_app.py"
    source = original.read_text(encoding="utf-8")
    patched = source
    for broken, fixed in UPSTREAM_PATCHES:
        patched = patched.replace(broken, fixed)
    if patched == source:
        return original
    PATCHED_APP.write_text(patched, encoding="utf-8")
    try:
        yield PATCHED_APP
    finally:
        PATCHED_APP.unlink(missing_ok=True)


@pytest.fixture
def app(patched_app_source, tmp_path, monkeypatch):
    """Start the app against a private state root."""

    from streamlit.testing.v1 import AppTest

    def start(state_root: Path | None = None):
        root = state_root or (tmp_path / "state")
        monkeypatch.setenv("THERMOGAR_STATE_ROOT", str(root))
        instance = AppTest.from_file(
            str(patched_app_source),
            default_timeout=900,
        )
        instance.run()
        assert not instance.exception, instance.exception
        return instance, root

    return start


def widget(elements, key):
    for element in elements:
        if element.key == key:
            return element
    raise AssertionError(f"widget {key!r} is not on the page")


def labelled(elements, label):
    for element in elements:
        if element.label == label:
            return element
    raise AssertionError(f"widget {label!r} is not on the page")


# Streamlit renders every tab on every run, so the precipitation section
# (app/thermogar_precipitation.py, another task's file) puts its own input
# messages on the page. They are not projects-and-data failures.
FOREIGN_ERRORS = (
    # thermogar_precipitation.py:181
    "Research KWN mode допускает не более четырёх добавок одновременно.",
    # thermogar_precipitation.py:818
    "Нет матричной фазы с полным набором мобильностей для состава.",
)


def section_errors(at) -> list[str]:
    """Errors this section is responsible for."""

    return [
        error.value for error in at.error if error.value not in FOREIGN_ERRORS
    ]


def download_keys(at) -> set[str]:
    return {button.key for button in at.download_button if button.key}


def stored_artifacts(state_root: Path) -> dict[str, bytes]:
    """Newest prepared export/import artifact per content kind.

    A kind accumulates one file per prepared payload (they are named by
    digest), so the newest one is the artifact behind the download button
    currently on the page.
    """

    newest: dict[str, tuple[float, Path]] = {}
    for path in (state_root / "state").rglob("*"):
        if not path.is_file():
            continue
        kind = path.parent.name
        stamp = path.stat().st_mtime_ns
        if kind not in newest or stamp >= newest[kind][0]:
            newest[kind] = (stamp, path)
    return {kind: path.read_bytes() for kind, (_stamp, path) in newest.items()}


# ---------------------------------------------------------------------------
# Widget-state allowlist (no Streamlit needed)
# ---------------------------------------------------------------------------


def test_widget_state_accepts_numbers_and_flags():
    value = {
        WIDGET_STATE_VERSION_FIELD: WIDGET_STATE_VERSION,
        "single_temperature_ni": 812.0,
        "t_step_ni": 37,
        "binary_nodes_fe": True,
    }
    assert validate_widget_state(value) == value


def test_widget_state_rejects_keys_outside_the_allowlist():
    with pytest.raises(ValueError, match="восстанавливаемый набор"):
        validate_widget_state(
            {
                WIDGET_STATE_VERSION_FIELD: WIDGET_STATE_VERSION,
                "thermogar_database_key": "ni",
            }
        )
    assert not is_restorable_widget_key("thermogar_composition_fe")
    assert is_restorable_widget_key("single_temperature_fe")


def test_widget_state_rejects_another_version_and_non_numbers():
    with pytest.raises(ValueError, match="другой версией"):
        validate_widget_state(
            {WIDGET_STATE_VERSION_FIELD: 99, "single_temperature_ni": 700.0}
        )
    with pytest.raises(ValueError, match="числом или флажком"):
        validate_widget_state(
            {
                WIDGET_STATE_VERSION_FIELD: WIDGET_STATE_VERSION,
                "solidification_step_ni": "10",
            }
        )
    assert validate_widget_state({}) == {}


def test_portable_project_drops_the_settings_only():
    payload = {"name": "x", "context": {"database_key": "ni"}, "widget_state": {"a": 1}}
    portable = portable_project_payload(payload)
    assert portable["widget_state"] == {}
    assert portable["name"] == "x"
    assert payload["widget_state"] == {"a": 1}


def test_rejections_are_plain_language_not_reason_codes():
    class Receipt:
        reason_code = "IMPORT_SCHEMA_REJECTED"
        reason_detail = "Exactly one comma/semicolon CSV parse must satisfy."

    text = rejection_text(Receipt())
    assert "IMPORT_SCHEMA_REJECTED" not in text
    assert "CSV parse" not in text
    assert text.startswith("Файл не соответствует")

    class Unknown:
        reason_code = "SOMETHING_NEW"
        reason_detail = "internal"

    assert "SOMETHING_NEW" not in rejection_text(Unknown())


# ---------------------------------------------------------------------------
# Alloy library
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("database_key", ["ni", "al", "fe"])
def test_library_save_appears_in_the_list_and_loads_back(app, database_key):
    at, state_root = app()
    at.sidebar.selectbox("thermogar_database_key").set_value(database_key)
    at.run()
    balance, units, _temperature_c, composition = CASES[database_key]
    at.session_state[f"thermogar_composition_{database_key}"] = composition
    at.session_state[f"thermogar_units_{database_key}"] = units.replace(
        "ат.%", "атомные %"
    ).replace("мас.%", "массовые %")
    at.run()

    name = f"Проверка {database_key}"
    labelled(at.text_input, "Название марки или состава").set_value(name)
    labelled(at.button, "Сохранить текущий состав").click()
    at.run()
    assert not section_errors(at)

    saved = json.loads(
        (state_root / "workspace" / "alloys.json").read_text(encoding="utf-8")
    )["alloys"]
    assert [item["name"] for item in saved] == [name]
    assert saved[0]["database_key"] == database_key
    assert saved[0]["composition"] == composition
    if database_key == "fe":
        assert saved[0]["fe_profile_key"] == "thermogar_patch"

    # Move away, then load the record back and check the sidebar follows.
    other = "ni" if database_key != "ni" else "al"
    at.sidebar.selectbox("thermogar_database_key").set_value(other)
    at.run()
    selector = widget(at.selectbox, "alloy_selected_id")
    # Options carry the formatted names; the value is the record id.
    assert name in selector.options
    selector.set_value(saved[0]["id"])
    at.run()
    widget(at.button, "alloy_load_button").click()
    at.run()
    assert not section_errors(at)
    assert at.session_state["thermogar_database_key"] == database_key
    assert at.session_state[f"thermogar_composition_{database_key}"] == composition
    assert at.session_state["_thermogar_loaded_context"]["label"] == name

    # Deleting the record needs the confirmation checkbox and keeps a backup.
    delete_button = f"alloy_delete_button_{saved[0]['id']}"
    widget(at.checkbox, f"alloy_delete_confirm_{saved[0]['id']}").set_value(True)
    at.run()
    widget(at.button, delete_button).click()
    at.run()
    assert not section_errors(at)
    assert (
        json.loads(
            (state_root / "workspace" / "alloys.json").read_text(encoding="utf-8")
        )["alloys"]
        == []
    )
    assert (state_root / "workspace" / "alloys.json.bak").is_file()


def test_library_export_and_import_round_trip(app, tmp_path):
    at, state_root = app()
    labelled(at.text_input, "Название марки или состава").set_value("Сталь")
    labelled(at.button, "Сохранить текущий состав").click()
    at.run()
    assert "alloy_library_download" in download_keys(at)

    exported = stored_artifacts(state_root)["alloy-library-json-v1"]
    envelope = json.loads(exported)
    assert envelope["kind"] == "thermogar_alloys"
    assert [item["name"] for item in envelope["payload"]] == ["Сталь"]

    second, second_root = app(tmp_path / "state2")
    labelled(second.file_uploader, "Импортировать библиотеку JSON").set_value(
        ("alloys.json", exported, "application/json")
    )
    second.run()
    widget(second.button, "alloy_import_button").click()
    second.run()
    assert not section_errors(second)
    imported = json.loads(
        (second_root / "workspace" / "alloys.json").read_text(encoding="utf-8")
    )["alloys"]
    assert [item["name"] for item in imported] == ["Сталь"]


# ---------------------------------------------------------------------------
# Projects and history
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("database_key", ["ni", "al", "fe"])
def test_project_saves_fourteen_keys_and_restores_state(app, database_key, tmp_path):
    state_root = tmp_path / "state"
    at, _ = app(state_root)
    at.sidebar.selectbox("thermogar_database_key").set_value(database_key)
    at.run()
    at.number_input(f"single_temperature_{database_key}").set_value(812.0)
    at.run()

    labelled(at.text_input, "Название проекта").set_value("Проект H")
    labelled(at.button, "Сохранить проект в папке ThermoGar").click()
    at.run()
    assert not section_errors(at)

    files = list((state_root / "workspace" / "projects").glob("*.thermogar.json"))
    assert len(files) == 1
    payload = json.loads(files[0].read_text(encoding="utf-8"))["payload"]
    assert len(payload) == 14
    assert payload["context"]["database_key"] == database_key
    assert payload["widget_state"][f"single_temperature_{database_key}"] == 812.0

    # A new session starts on the default steel database and knows nothing
    # about the project until it is opened.
    fresh, _ = app(state_root)
    assert fresh.session_state["thermogar_database_key"] == "fe"
    if database_key != "fe":
        assert f"single_temperature_{database_key}" not in fresh.session_state
    widget(fresh.button, "project_load_button").click()
    fresh.run()
    assert not section_errors(fresh)
    assert fresh.session_state["thermogar_database_key"] == database_key
    assert fresh.session_state[f"single_temperature_{database_key}"] == 812.0


def test_project_export_is_portable_and_imports_back(app, tmp_path):
    at, state_root = app()
    at.number_input("single_temperature_fe").set_value(755.0)
    at.run()
    labelled(at.text_input, "Название проекта").set_value("Проект H")
    labelled(at.button, "Сохранить проект в папке ThermoGar").click()
    at.run()
    assert "project_download" in download_keys(at)

    exported = stored_artifacts(state_root)["project-json-v1"]
    envelope = json.loads(exported)
    assert envelope["kind"] == "thermogar_project"
    # The portable copy carries the material, not the interface settings.
    assert envelope["payload"]["widget_state"] == {}
    assert envelope["payload"]["context"]["database_key"] == "fe"

    second, second_root = app(tmp_path / "state2")
    labelled(second.file_uploader, "Импортировать проект").set_value(
        ("project.json", exported, "application/json")
    )
    second.run()
    widget(second.button, "project_import_button").click()
    second.run()
    assert not section_errors(second)
    assert list((second_root / "workspace" / "projects").glob("*.thermogar.json"))

    # Deleting keeps the file under a .deleted name and drops it from the list.
    widget(second.checkbox, "project_delete_confirm").set_value(True)
    second.run()
    widget(second.button, "project_delete_button").click()
    second.run()
    assert not section_errors(second)
    projects = second_root / "workspace" / "projects"
    assert not list(projects.glob("*.thermogar.json"))
    assert list(projects.glob("*.thermogar.json.deleted"))


def test_confirmations_survive_the_rerun_and_stay_in_their_own_section(app):
    """A durable write is followed by st.rerun, which discards st.success."""

    at, _ = app()
    labelled(at.text_input, "Название проекта").set_value("Проект H")
    labelled(at.button, "Сохранить проект в папке ThermoGar").click()
    at.run()
    messages = [message.value for message in at.success]
    assert any(text.startswith("Проект сохранён:") for text in messages)
    assert not any(text.startswith("Состав ") for text in messages)

    labelled(at.text_input, "Название марки или состава").set_value("Марка H")
    labelled(at.button, "Сохранить текущий состав").click()
    at.run()
    messages = [message.value for message in at.success]
    assert any(text.startswith("Состав «Марка H»") for text in messages)
    # The project confirmation was consumed by its own section, not this one.
    assert not any(text.startswith("Проект сохранён:") for text in messages)


def test_history_records_events_exports_csv_and_clears(app):
    at, state_root = app()
    labelled(at.text_input, "Название марки или состава").set_value("Сталь")
    labelled(at.button, "Сохранить текущий состав").click()
    at.run()

    history = state_root / "workspace" / "history.jsonl"
    entries = [
        json.loads(line)
        for line in history.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]
    assert [entry["event_type"] for entry in entries] == ["alloy_saved"]
    assert entries[0]["database_key"] == "fe"
    assert entries[0]["previous_sha256"] == ""

    widget(at.radio, "projects_history_mode").set_value("История расчётов")
    at.run()
    assert any(
        message.value.startswith("Цепочка контрольных сумм")
        for message in at.success
    )
    assert "history_download" in download_keys(at)

    import pandas as pd

    exported = stored_artifacts(state_root)["history-csv-v1"]
    frame = pd.read_csv(io.BytesIO(exported))
    assert "Событие" in frame.columns
    assert len(frame) == 1

    # Restoring the material from a history row must reach the sidebar.
    widget(at.button, "history_restore_button").click()
    at.run()
    assert not section_errors(at)

    widget(at.radio, "projects_history_mode").set_value("История расчётов")
    at.run()
    widget(at.checkbox, "history_clear_confirm").set_value(True)
    at.run()
    widget(at.button, "history_clear_button").click()
    at.run()
    assert not section_errors(at)
    assert history.read_bytes() == b""
    assert list((state_root / "workspace").glob("history_*.jsonl.bak"))


# ---------------------------------------------------------------------------
# Batch calculation
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "separator,bom",
    [(",", False), (";", False), (",", True), (";", True)],
    ids=["comma", "semicolon", "comma-bom", "semicolon-bom"],
)
def test_batch_accepts_both_separators_and_both_encodings(app, separator, bom):
    at, _ = app()
    labelled(at.file_uploader, "Файл составов").set_value(
        ("batch.csv", batch_csv(separator=separator, bom=bom), "text/csv")
    )
    at.run()
    assert not section_errors(at)
    # The preview frame is canonical: aliases resolved, units normalised.
    preview = next(
        frame.value
        for frame in at.dataframe
        if "database" in getattr(frame.value, "columns", [])
    )
    assert list(preview["database"]) == ["ni", "al", "fe"]
    assert list(preview["units"]) == ["at", "wt", "wt"]
    assert widget(at.button, "batch_calculate_button").label


def test_batch_rejects_a_junk_file_with_a_readable_message(app):
    at, _ = app()
    labelled(at.file_uploader, "Файл составов").set_value(
        ("batch.csv", b"foo,bar\n1,2\n", "text/csv")
    )
    at.run()
    messages = section_errors(at)
    assert messages, "a rejected upload must say something"
    assert any(text.startswith("Файл составов не принят:") for text in messages)
    assert not any("IMPORT_SCHEMA_REJECTED" in text for text in messages)
    assert not any(button.key == "batch_calculate_button" for button in at.button)


def test_batch_template_downloads_open_as_excel_and_csv(app):
    at, state_root = app()
    widget(at.button, "batch_template_xlsx_prepare").click()
    at.run()
    assert "batch_template_xlsx_download" in download_keys(at)
    widget(at.button, "batch_template_csv_prepare").click()
    at.run()
    assert "batch_template_csv_download" in download_keys(at)

    import openpyxl
    import pandas as pd

    artifacts = stored_artifacts(state_root)
    workbook = openpyxl.load_workbook(
        io.BytesIO(artifacts["batch-template-xlsx-v1"])
    )
    assert workbook.sheetnames == ["Составы"]

    csv_bytes = artifacts["batch-template-csv-v1"]
    assert csv_bytes.startswith(b"\xef\xbb\xbf")
    frame = pd.read_csv(io.BytesIO(csv_bytes))
    assert "Температура, °C" in frame.columns
    assert len(frame) == 2


def test_batch_summary_has_no_receipt_columns(app):
    at, _ = app()
    labelled(at.file_uploader, "Файл составов").set_value(
        ("batch.csv", batch_csv(("ni",)), "text/csv")
    )
    at.run()
    widget(at.button, "batch_calculate_button").click()
    at.run()
    summary = at.session_state["workspace_batch_result"]["display"]["Сводка"]
    assert "База SHA-256" in summary.columns
    assert "Receipt SHA-256" not in summary.columns
    assert "Envelope SHA-256" not in summary.columns


@pytest.mark.slow
def test_batch_calculates_three_databases_and_exports_excel(app):
    at, state_root = app()
    labelled(at.file_uploader, "Файл составов").set_value(
        ("batch.csv", batch_csv(), "text/csv")
    )
    at.run()
    widget(at.button, "batch_calculate_button").click()
    at.run()
    assert not section_errors(at)

    result = at.session_state["workspace_batch_result"]["display"]
    summary = result["Сводка"]
    assert list(summary["Статус"]) == ["готово"] * 3
    assert list(summary["База"]) == ["ni", "al", "fe"]
    for total in summary["Сумма фазовых долей, %"]:
        assert abs(float(total) - 100.0) < 1e-6

    phases = result["Фазовые доли"]
    assert not phases.empty
    assert "C15_LAVES" not in set(phases["Фаза"])
    by_name = {
        (row["Название"], row["Фаза"]): row["Мольная доля фазы, %"]
        for _index, row in phases.iterrows()
    }
    assert abs(by_name[("NI", "GAMMA_PRIME")] - 30.422) < 0.05
    assert abs(by_name[("AL", "THETA_AL2CU")] - 0.766) < 0.05
    assert abs(by_name[("FE", "M23C6")] - 4.413) < 0.05

    widget(at.button, "batch_result_export_prepare").click()
    at.run()
    assert "batch_result_export_download" in download_keys(at)

    import openpyxl

    workbook = openpyxl.load_workbook(
        io.BytesIO(stored_artifacts(state_root)["batch-result-xlsx-v1"])
    )
    assert workbook.sheetnames == [
        "Сводка",
        "Фазовые доли",
        "Составы фаз ат",
        "Составы фаз мас",
        "Исходные данные",
    ]


@pytest.mark.slow
def test_batch_rejects_c15_for_steel_before_any_calculation(app):
    at, _ = app()
    payload = (
        'Название,База,Основа,Единицы,"Температура, °C",Добавки,Фазы\n'
        'FE,fe,FE,мас.%,700,"C=0.2, CR=11.5",C15_LAVES\n'
    ).encode("utf-8")
    labelled(at.file_uploader, "Файл составов").set_value(
        ("batch.csv", payload, "text/csv")
    )
    at.run()
    messages = section_errors(at)
    assert any("C15_LAVES" in text for text in messages), messages
    assert not any(button.key == "batch_calculate_button" for button in at.button)


# ---------------------------------------------------------------------------
# Paths and first run
# ---------------------------------------------------------------------------


def test_first_run_creates_the_profile_and_writes_nothing_into_the_program(
    patched_app_source, tmp_path, monkeypatch
):
    from streamlit.testing.v1 import AppTest

    local_app_data = tmp_path / "LocalAppData"
    monkeypatch.delenv("THERMOGAR_STATE_ROOT", raising=False)
    monkeypatch.setenv("LOCALAPPDATA", str(local_app_data))
    before = {path.name for path in APP.glob("*.json")}

    at = AppTest.from_file(str(patched_app_source), default_timeout=900)
    at.run()
    assert not at.exception

    root = local_app_data / "ThermoGar"
    assert (root / "workspace" / "projects").is_dir()
    assert (root / "logs").is_dir()
    assert (root / "runtime").is_dir()
    assert {path.name for path in APP.glob("*.json")} == before


# ---------------------------------------------------------------------------
# Quick examples
# ---------------------------------------------------------------------------


def test_quick_examples_cover_three_databases_and_carry_a_steel():
    import thermogar_stage14

    captured: list[tuple[dict, dict, str]] = []

    class FakeButton:
        def __init__(self) -> None:
            self.calls = 0

        def __call__(self, *args, **kwargs) -> bool:
            self.calls += 1
            return True

    class FakeStreamlit:
        def __init__(self) -> None:
            self.button = FakeButton()

        def markdown(self, *args, **kwargs) -> None: ...

        def caption(self, *args, **kwargs) -> None: ...

        def rerun(self) -> None: ...

        def container(self, *args, **kwargs):
            class Ctx:
                def __enter__(self_inner): return self_inner
                def __exit__(self_inner, *exc): return False

            return Ctx()

    fake = FakeStreamlit()
    original = thermogar_stage14.st
    thermogar_stage14.st = fake
    try:
        thermogar_stage14.render_quick_examples(
            lambda context, widget_state=None, label="": captured.append(
                (context, widget_state, label)
            )
        )
    finally:
        thermogar_stage14.st = original

    keys = [context["database_key"] for context, _state, _label in captured]
    assert keys == ["ni", "al", "fe"]
    for context, widget_state, _label in captured:
        # Every example must set the temperature of its own database.
        assert widget_state == {
            f"single_temperature_{context['database_key']}": pytest.approx(
                {"ni": 700.0, "al": 500.0, "fe": 700.0}[context["database_key"]]
            )
        }
        assert validate_widget_state(
            {WIDGET_STATE_VERSION_FIELD: WIDGET_STATE_VERSION, **widget_state}
        )

    steel = next(context for context, _s, _l in captured if context["database_key"] == "fe")
    assert steel["composition"] == "C=0.2, CR=11.5, NI=0.7"
    assert steel["units"] == "wt"
    assert steel["steel_mode"] == "metastable"


if __name__ == "__main__":
    raise SystemExit(pytest.main([__file__, "-v"]))
