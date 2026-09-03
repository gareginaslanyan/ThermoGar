"""Матрица разделов 3F через streamlit.testing.v1.AppTest.

Пять разделов приложения — «Расчёты», «Диаграммы», «Затвердевание»,
«Энергии», «Свойства» — проверяются на трёх базах (ni/al/fe). Для каждой
ячейки: нажатие кнопки расчёта, отсутствие ``at.exception`` и неожиданных
``st.error``, осмысленный результат в session_state и проверка выгрузок
**по байтам**: Excel открывается ``openpyxl``, PNG начинается с
``\\x89PNG``, CSV разбирается ``pandas``.

Байты выгрузок берутся перехватом ``st.download_button``: AppTest отдаёт
только mock-URL, поэтому обёртка запоминает то, что приложение реально
передало кнопке.

Запуск:

    python -m pytest tools/test_ui_f.py -m "not slow"
    python -m pytest tools/test_ui_f.py -m slow
"""

from __future__ import annotations

import io
import sys
from pathlib import Path
from typing import Any, Callable, Iterator

import pandas as pd
import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent
APP_PATH = PROJECT_ROOT / "app" / "ThermoGar_app.py"
if str(PROJECT_ROOT / "app") not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT / "app"))

import streamlit as st  # noqa: E402
from streamlit.testing.v1 import AppTest  # noqa: E402

# Ошибки чужих модулей кинетики (зона 3G), которые появляются на первом
# рендере и не относятся к разделам 3F. Они вынесены в отчёт, а не чинятся.
KNOWN_FOREIGN_ERRORS = (
    "Research KWN mode допускает не более четырёх добавок",
    "Нет матричной фазы с полным набором мобильностей",
)

BASES: dict[str, dict[str, Any]] = {
    "ni": {
        "composition": "AL=15",
        "units": "атомные %",
        "balance": "NI",
        "temperature": 700.0,
        "scan": (600.0, 1000.0, 100.0),
        "variable": "AL",
        "concentration": (5.0, 25.0, 5.0),
        "solidification_start": 1500.0,
        "solidification_step": 25.0,
    },
    "al": {
        "composition": "CU=4, MG=1",
        "units": "массовые %",
        "balance": "AL",
        "temperature": 500.0,
        "scan": (300.0, 600.0, 75.0),
        "variable": "CU",
        "concentration": (1.0, 5.0, 1.0),
        "solidification_start": 700.0,
        "solidification_step": 25.0,
    },
    "fe": {
        "composition": "C=0.2, CR=11.5, NI=0.7",
        "units": "массовые %",
        "balance": "FE",
        "temperature": 700.0,
        "scan": (500.0, 900.0, 100.0),
        "variable": "C",
        "concentration": (0.1, 0.5, 0.1),
        "solidification_start": 1500.0,
        "solidification_step": 25.0,
    },
}

DATABASES = tuple(BASES)


# ---------------------------------------------------------------------------
# Запуск приложения и перехват выгрузок
# ---------------------------------------------------------------------------


class Session:
    """Один прогон приложения вместе с байтами всех кнопок скачивания."""

    def __init__(self, app_test: AppTest, downloads: dict[str, bytes]) -> None:
        self.at = app_test
        self.downloads = downloads

    def click(self, key: str) -> "Session":
        self.at.button(key=key).click().run()
        return self

    def state(self, key: str) -> Any:
        assert key in self.at.session_state, f"нет результата в {key}"
        return self.at.session_state[key]

    def unexpected_errors(self) -> list[str]:
        return [
            element.value
            for element in self.at.error
            if not any(known in element.value for known in KNOWN_FOREIGN_ERRORS)
        ]

    def assert_clean(self) -> None:
        assert not self.at.exception, [str(e.value) for e in self.at.exception]
        assert not self.unexpected_errors(), self.unexpected_errors()

    def download(self, file_name: str) -> bytes:
        assert file_name in self.downloads, (
            f"кнопка скачивания {file_name!r} не найдена: "
            f"{sorted(self.downloads)}"
        )
        return self.downloads[file_name]


ELASTIC_ROW_VALUES = {
    "young_gpa": 200.0,
    "poisson": 0.3,
    "origin": "measured",
    "source": "учебное значение для проверки интерфейса",
    "reference_temperature_c": 25.0,
}


def start(
    database_key: str,
    *,
    session: dict[str, Any] | None = None,
    prepare: Callable[[AppTest], None] | None = None,
    fill_elastic_editor: bool = False,
) -> Session:
    """Запустить приложение на выбранной базе с составом из COMMON."""

    profile = BASES[database_key]
    downloads: dict[str, bytes] = {}
    original = st.download_button

    if fill_elastic_editor:
        # AppTest не умеет вводить значения в st.data_editor, поэтому
        # редактор упругих свойств подменяется: приложение получает те же
        # строки с заполненными E, ν и происхождением.
        original_editor = st.data_editor

        def filled_editor(frame: Any, *args: Any, **kwargs: Any) -> Any:
            original_editor(frame, *args, **kwargs)
            edited = frame.copy()
            for column, value in ELASTIC_ROW_VALUES.items():
                edited[column] = value
            return edited

        st.data_editor = filled_editor

    def capture(label: Any, data: Any = None, *args: Any, **kwargs: Any) -> Any:
        name = str(kwargs.get("file_name") or label)
        if isinstance(data, (bytes, bytearray)):
            downloads[name] = bytes(data)
        elif isinstance(data, str):
            downloads[name] = data.encode("utf-8")
        return original(label, data, *args, **kwargs)

    st.download_button = capture
    app_test = AppTest.from_file(str(APP_PATH), default_timeout=1800)
    app_test.session_state["thermogar_database_key"] = database_key
    app_test.session_state[f"thermogar_composition_{database_key}"] = profile[
        "composition"
    ]
    app_test.session_state[f"thermogar_units_{database_key}"] = profile["units"]
    app_test.session_state[f"thermogar_balance_{database_key}"] = profile["balance"]
    for name, value in (session or {}).items():
        app_test.session_state[name] = value
    app_test.run()
    if prepare is not None:
        prepare(app_test)
        app_test.run()
    return Session(app_test, downloads)


@pytest.fixture(autouse=True)
def _private_state_root(tmp_path, monkeypatch) -> None:
    """Приватный профиль на тест: не писать в общий %LOCALAPPDATA%\\ThermoGar."""

    monkeypatch.setenv("THERMOGAR_STATE_ROOT", str(tmp_path / "state"))


@pytest.fixture(autouse=True)
def _restore_streamlit_seams() -> Iterator[None]:
    download_button = st.download_button
    data_editor = st.data_editor
    yield
    st.download_button = download_button
    st.data_editor = data_editor


def set_number(app_test: AppTest, label: str, value: float) -> None:
    """Задать значение number_input по подписи (у виджета нет своего key)."""

    for widget in app_test.number_input:
        if widget.label == label:
            widget.set_value(float(value))
            return
    raise AssertionError(f"не найдено поле {label!r}")


def set_select(app_test: AppTest, label: str, value: str) -> None:
    for widget in app_test.selectbox:
        if widget.label == label:
            widget.set_value(value)
            return
    raise AssertionError(f"не найден список {label!r}")


# ---------------------------------------------------------------------------
# Проверки выгрузок по байтам
# ---------------------------------------------------------------------------


def assert_xlsx(payload: bytes, *, sheets: int = 1) -> list[str]:
    import openpyxl

    assert payload[:2] == b"PK", payload[:8]
    book = openpyxl.load_workbook(io.BytesIO(payload))
    assert len(book.sheetnames) >= sheets, book.sheetnames
    return list(book.sheetnames)


def assert_png(payload: bytes) -> None:
    assert payload.startswith(b"\x89PNG\r\n\x1a\n"), payload[:8]
    assert len(payload) > 1024


def assert_csv(payload: bytes) -> pd.DataFrame:
    frame = pd.read_csv(io.BytesIO(payload))
    assert not frame.empty
    return frame


def assert_fractions_close_to_one(frame: pd.DataFrame, column: str) -> None:
    total = float(frame[column].sum())
    assert abs(total - 100.0) < 0.5, total


# ---------------------------------------------------------------------------
# Старт приложения
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("database_key", DATABASES)
def test_startup_is_clean(database_key: str) -> None:
    """Первый рендер: без исключений и без ложного BINDING_STALE."""

    session = start(database_key)
    session.assert_clean()
    assert not any(
        "BINDING_STALE" in element.value for element in session.at.error
    )


# ---------------------------------------------------------------------------
# Раздел «Расчёты»
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("database_key", DATABASES)
def test_single_equilibrium(database_key: str) -> None:
    profile = BASES[database_key]
    session = start(
        database_key,
        session={
            f"single_temperature_{database_key}": profile["temperature"],
        },
    )
    session.click("single_calculate")
    session.assert_clean()

    display = session.state("_thermogar_vlb_b3_result_equilibrium_single")["display"]
    summary = display["summary"]
    assert not summary.empty
    assert_fractions_close_to_one(summary, "Мольная доля фазы, %")
    assert "C15_LAVES" not in set(summary["Фаза"])
    # Fe раньше шёл отдельным маршрутом без таблиц состава фаз.
    assert not display["phase_at"].empty
    assert not display["phase_wt"].empty
    assert len(display["phase_at"]) == len(summary)

    sheets = assert_xlsx(
        session.download("ThermoGar_equilibrium.xlsx"), sheets=6
    )
    assert "Составы фаз ат" in sheets


@pytest.mark.parametrize("database_key", DATABASES)
def test_temperature_scan(database_key: str) -> None:
    profile = BASES[database_key]
    t_min, t_max, t_step = profile["scan"]
    session = start(
        database_key,
        session={
            f"t_min_{database_key}": t_min,
            f"t_max_{database_key}": t_max,
            f"t_step_{database_key}": t_step,
        },
    )
    session.click("temperature_calculate")
    session.assert_clean()

    display = session.state(
        "_thermogar_vlb_b3_result_equilibrium_temperature_scan"
    )["display"]
    data = display["data"]
    assert len(data) == 5
    assert "C15_LAVES" not in data.columns
    for _index, row in data.iterrows():
        total = float(row.drop(labels=["Температура, °C"]).sum())
        assert abs(total - 100.0) < 0.5, total

    assert_xlsx(session.download("ThermoGar_temperature_scan.xlsx"), sheets=3)
    assert_csv(session.download("ThermoGar_temperature_scan.csv"))
    assert_png(session.download("ThermoGar_temperature_scan.png"))


@pytest.mark.parametrize("database_key", DATABASES)
def test_concentration_scan(database_key: str) -> None:
    profile = BASES[database_key]
    c_min, c_max, c_step = profile["concentration"]
    variable = profile["variable"]

    def prepare(app_test: AppTest) -> None:
        set_select(app_test, "Изменяемый элемент", variable)

    session = start(database_key, prepare=prepare)
    set_number(session.at, f"{variable}: от, %", c_min)
    set_number(session.at, f"{variable}: до, %", c_max)
    set_number(session.at, f"{variable}: шаг, %", c_step)
    session.at.run()

    session.click("concentration_calculate")
    session.assert_clean()

    display = session.state(
        "_thermogar_vlb_b3_result_equilibrium_composition_scan"
    )["display"]
    data = display["data"]
    assert len(data) == 5
    assert "C15_LAVES" not in data.columns

    assert_xlsx(session.download("ThermoGar_concentration_scan.xlsx"), sheets=3)
    assert_csv(session.download("ThermoGar_concentration_scan.csv"))
    assert_png(session.download("ThermoGar_concentration_scan.png"))


# ---------------------------------------------------------------------------
# Раздел «Затвердевание»
# ---------------------------------------------------------------------------


SOLIDIFICATION_METHODS = (
    "Сравнить равновесное и Scheil–Gulliver",
    "Только равновесное затвердевание",
    "Только Scheil–Gulliver",
)


@pytest.mark.slow
@pytest.mark.parametrize("database_key", DATABASES)
@pytest.mark.parametrize("method", SOLIDIFICATION_METHODS)
def test_solidification(database_key: str, method: str) -> None:
    profile = BASES[database_key]
    start_key = (
        f"solidification_start_13_1_{database_key}_thermogar_patch"
        if database_key == "fe"
        else f"solidification_start_13_1_{database_key}"
    )
    session = start(
        database_key,
        session={
            f"solidification_method_{database_key}": method,
            start_key: profile["solidification_start"],
            f"solidification_step_{database_key}": profile["solidification_step"],
        },
    )
    session.click("solidification_calculate")
    session.assert_clean()

    state = session.state("solidification_result")
    assert state["database_key"] == database_key
    results = state["results"]
    expected = {
        "Сравнить равновесное и Scheil–Gulliver": {"equilibrium", "scheil"},
        "Только равновесное затвердевание": {"equilibrium"},
        "Только Scheil–Gulliver": {"scheil"},
    }[method]
    assert set(results) == expected
    assert not state["summary"].empty

    assert_xlsx(session.download("ThermoGar_solidification.xlsx"), sheets=2)
    assert_png(session.download("ThermoGar_liquid_fraction.png"))
    archive = session.download("ThermoGar_solidification_results.zip")
    assert archive[:2] == b"PK"


# ---------------------------------------------------------------------------
# Раздел «Энергии»
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("database_key", DATABASES)
def test_energy_curve(database_key: str) -> None:
    session = start(
        database_key,
        session={
            f"energy_t_min_{database_key}": 600.0,
            f"energy_t_max_{database_key}": 1000.0,
            f"energy_t_step_{database_key}": 100.0,
        },
    )
    session.click("energy_curve_calculate")
    session.assert_clean()

    state = session.state("energy_curve_result")
    assert state["database_key"] == database_key
    assert not state["absolute"].empty
    assert not state["relative"].empty

    assert_xlsx(session.download("ThermoGar_phase_energies.xlsx"), sheets=2)
    assert_png(session.download("ThermoGar_phase_energies.png"))


@pytest.mark.parametrize("database_key", DATABASES)
def test_driving_force(database_key: str) -> None:
    session = start(
        database_key,
        session={
            f"driving_t_min_{database_key}": 600.0,
            f"driving_t_max_{database_key}": 1000.0,
            f"driving_t_step_{database_key}": 100.0,
        },
    )
    session.click("driving_force_calculate")
    session.assert_clean()

    state = session.state("driving_force_result")
    assert state["database_key"] == database_key
    assert not state["data"].empty

    assert_xlsx(session.download("ThermoGar_driving_force.xlsx"), sheets=2)
    assert_png(session.download("ThermoGar_driving_force.png"))


TZERO_WINDOWS = {
    "ni": (1230.0, 1630.0),
    "al": (530.0, 830.0),
    "fe": (600.0, 760.0),
}


@pytest.mark.parametrize("database_key", DATABASES)
def test_tzero_in_narrow_window(database_key: str) -> None:
    """T₀ ищется в узком окне вокруг перехода; UI это объясняет."""

    t_min, t_max = TZERO_WINDOWS[database_key]
    session = start(
        database_key,
        session={
            f"tzero_t_min_{database_key}": t_min,
            f"tzero_t_max_{database_key}": t_max,
        },
    )
    session.click("tzero_calculate")
    session.assert_clean()

    state = session.state("tzero_result")
    assert state["database_key"] == database_key
    data = state["data"]
    assert not data.empty
    assert int(data["Решение найдено"].sum()) > 0

    captions = " ".join(element.value for element in session.at.caption)
    assert "узким" in captions

    assert_xlsx(session.download("ThermoGar_T0.xlsx"), sheets=2)
    assert_png(session.download("ThermoGar_T0.png"))


# ---------------------------------------------------------------------------
# Раздел «Свойства»
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("database_key", DATABASES)
def test_density_single(database_key: str) -> None:
    profile = BASES[database_key]
    session = start(
        database_key,
        session={f"physical_temperature_{database_key}": profile["temperature"]},
    )
    session.click("physical_single_calculate")
    session.assert_clean()

    state = session.state(
        "_thermogar_vlb_b4b_result_property_density_single"
    )
    projection = state["projections"][0]
    assert projection["mass_coverage_pct"] > 0.0
    assert projection["phase_rows"]
    if projection["alloy_density_kg_m3"] is None:
        # Al: в physical_data_v103.pdb нет модели плотности THETA_AL2CU.
        texts = " ".join(element.value for element in session.at.info)
        assert "покрытие физической базы" in texts
        assert projection["mass_coverage_pct"] < 100.0
    else:
        assert projection["alloy_density_kg_m3"] > 1000.0

    sheets = assert_xlsx(
        session.download("ThermoGar_density_single.xlsx"), sheets=3
    )
    assert "Параметры" in sheets


@pytest.mark.parametrize("database_key", DATABASES)
def test_density_temperature_scan(database_key: str) -> None:
    profile = BASES[database_key]
    t_min, t_max, t_step = profile["scan"]
    session = start(
        database_key,
        session={
            f"physical_t_min_{database_key}": t_min,
            f"physical_t_max_{database_key}": t_min + 2.0 * t_step,
            f"physical_t_step_{database_key}": t_step,
        },
    )
    session.click("physical_scan_calculate")
    session.assert_clean()

    state = session.state(
        "_thermogar_vlb_b4b_result_property_density_temperature"
    )
    assert len(state["projections"]) == 3

    assert_xlsx(session.download("ThermoGar_density_scan.xlsx"), sheets=2)
    assert_png(session.download("ThermoGar_density_scan.png"))


@pytest.mark.parametrize("database_key", DATABASES)
def test_elastic_vrh(database_key: str) -> None:
    profile = BASES[database_key]
    session = start(
        database_key,
        session={
            f"b4b2_elastic_temperature_{database_key}": profile["temperature"],
        },
        fill_elastic_editor=True,
    )
    session.click("b4b2_elastic_prepare_calculate")
    session.assert_clean()
    assert "_thermogar_vlb_b4b_result_property_elastic_prepare" in (
        session.at.session_state
    )

    session.click("b4b2_elastic_vrh_calculate")
    session.assert_clean()
    state = session.state("_thermogar_vlb_b4b_result_property_elastic_vrh")
    summary = state["projection"]["summary"]
    assert summary["E_Hill_GPa"] > 0.0
    assert summary["G_Hill_GPa"] > 0.0
    assert summary["E_Reuss_GPa"] <= summary["E_Hill_GPa"] <= summary["E_Voigt_GPa"]

    sheets = assert_xlsx(
        session.download("ThermoGar_elastic_vrh.xlsx"), sheets=3
    )
    assert "Voigt-Reuss-Hill" in sheets


@pytest.mark.parametrize("database_key", DATABASES)
def test_strengthening(database_key: str) -> None:
    session = start(
        database_key,
        session={
            f"b4b2_strengthening_rule_{database_key}": "Линейная сумма",
            f"b4b2_strengthening_confirmation_{database_key}": True,
            f"b4b2_strengthening_provenance_{database_key}": "учебные значения",
            f"b4b2_hall_use_{database_key}": True,
            f"b4b2_taylor_use_{database_key}": True,
        },
    )
    session.click("b4b2_strengthening_calculate")
    session.assert_clean()

    state = session.state("_thermogar_vlb_b4b_result_property_strengthening")
    projection = state["projection"]
    assert projection["contribution_rows"]
    assert projection["total_mpa"] is not None and projection["total_mpa"] > 0.0

    assert_xlsx(session.download("ThermoGar_strengthening.xlsx"), sheets=2)


# ---------------------------------------------------------------------------
# Раздел «Диаграммы» — только на грубой сетке по умолчанию
# ---------------------------------------------------------------------------


@pytest.mark.slow
@pytest.mark.parametrize("database_key", DATABASES)
def test_binary_diagram(database_key: str) -> None:
    session = start(database_key)
    session.click("binary_calculate")
    session.assert_clean()

    result = session.state(f"binary_result_{database_key}")
    assert result["figure"] is not None
    assert not result["boundaries"].empty

    assert_xlsx(session.download("ThermoGar_binary_diagram.xlsx"), sheets=2)
    assert_png(session.download("ThermoGar_binary_diagram.png"))


@pytest.mark.slow
@pytest.mark.parametrize("database_key", DATABASES)
def test_isopleth_diagram(database_key: str) -> None:
    session = start(database_key)
    session.click("isopleth_calculate")
    session.assert_clean()

    result = session.state(f"isopleth_result_{database_key}")
    assert result["figure"] is not None

    assert_xlsx(
        session.download("ThermoGar_multicomponent_section.xlsx"), sheets=2
    )
    assert_png(session.download("ThermoGar_multicomponent_section.png"))


@pytest.mark.slow
@pytest.mark.parametrize("database_key", DATABASES)
def test_ternary_diagram(database_key: str) -> None:
    session = start(database_key)
    session.click("ternary_calculate")
    session.assert_clean()

    result = session.state(f"ternary_result_{database_key}")
    assert result["figure"] is not None

    assert_xlsx(session.download("ThermoGar_ternary_diagram.xlsx"), sheets=2)
    assert_png(session.download("ThermoGar_ternary_diagram.png"))


@pytest.mark.slow
@pytest.mark.parametrize("database_key", DATABASES)
def test_ternary_phase_map(database_key: str) -> None:
    session = start(database_key)
    session.click("ternary_map_calculate")
    session.assert_clean()

    result = session.state(f"ternary_map_result_{database_key}")
    data = result["data"]
    assert not data.empty
    assert (data["Статус"] == "рассчитано").any()

    assert_xlsx(session.download("ThermoGar_ternary_phase_map.xlsx"), sheets=2)
    assert_png(session.download("ThermoGar_ternary_phase_map.png"))
    assert_csv(session.download("ThermoGar_ternary_phase_map.csv"))


@pytest.mark.parametrize("database_key", DATABASES)
def test_phase_map_needs_three_elements(database_key: str) -> None:
    """Карта строится ровно по трём элементам: UI не даёт выбрать иначе."""

    session = start(database_key)
    chosen = [
        widget.value
        for widget in session.at.selectbox
        if str(widget.key or "").startswith(
            ("ternary_map_x_", "ternary_map_y_", "ternary_map_dependent_")
        )
    ]
    # Карта берёт ровно три элемента: два по осям и один как остаток.
    # Списки B и C формируются из уже не занятых элементов, поэтому
    # повторить элемент невозможно, а четвёртой оси в интерфейсе нет.
    assert len(chosen) == 3, chosen
    assert len(set(chosen)) == 3, chosen


# ---------------------------------------------------------------------------
# Смена базы не должна показывать чужой результат
# ---------------------------------------------------------------------------


def test_results_do_not_survive_a_database_change() -> None:
    session = start(
        "ni",
        session={"single_temperature_ni": BASES["ni"]["temperature"]},
    )
    session.click("single_calculate")
    session.assert_clean()
    assert "_thermogar_vlb_b3_result_equilibrium_single" in session.at.session_state

    session.at.session_state["thermogar_database_key"] = "fe"
    session.at.session_state["thermogar_composition_fe"] = BASES["fe"]["composition"]
    session.at.session_state["thermogar_units_fe"] = BASES["fe"]["units"]
    session.at.run()

    assert not session.at.exception
    assert "_thermogar_vlb_b3_result_equilibrium_single" not in (
        session.at.session_state
    )
