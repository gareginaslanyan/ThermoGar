"""Verified, path-free state-transfer boundary for VLB Wave B4A.

The public UI seam carries only prepared requests and scalar widget metadata.
Uploaded bytes, persisted paths, and download bytes remain owned by
``StateStore``.  The module adds no verified-loader schema, feature ID, reason,
or Paths API.
"""

from __future__ import annotations

from dataclasses import dataclass, fields
from datetime import datetime, timezone
from io import BytesIO, StringIO
from pathlib import Path
import csv
import hashlib
import json
import math
import re
from typing import Any, Callable, Mapping, Sequence
from zipfile import BadZipFile, ZipFile

import thermogar_verified_loaders as verified_loaders
from thermogar_release_policy import (
    APP_STAGE,
    APP_VERSION,
    PRODUCTION_USE,
    RELEASE_CLASS,
    SCIENTIFIC_MATERIAL_STATUS,
    SOFTWARE_RELEASE_STATUS,
)
from thermogar_secure_io import (
    MAX_WORKSPACE_FILE_BYTES,
    atomic_write_bytes,
    ensure_plain_directory,
    held_verified_snapshot,
    read_verified_snapshot,
)


ALLOY_UI_TYPES = ("json",)
PROJECT_UI_TYPES = ("json",)
BATCH_UI_TYPES = ("csv", "xlsx")
ALLOY_UPLOAD_KEY = "alloy_library_uploader"
PROJECT_UPLOAD_KEY = "project_uploader"
BATCH_UPLOAD_KEY = "batch_file_uploader"

ENVELOPE_KEYS = ("schema_version", "kind", "exported_at", "payload", "sha256")
PROJECT_PAYLOAD_KEYS = (
    "schema_version", "kind", "name", "description", "created_at",
    "updated_at", "app_stage", "app_version", "release_class",
    "software_release_status", "scientific_material_status", "production_use",
    "context", "widget_state",
)
HISTORY_HEADERS = (
    "Время", "Событие", "Тип", "База", "Основа", "Единицы", "Состав",
    "Подробности", "База SHA-256", "Запись SHA-256",
)
BATCH_REQUIRED_HEADERS = ("name", "database", "balance", "units", "temperature_C")
BATCH_OPTIONAL_HEADERS = ("pressure_Pa", "steel_mode", "phases", "composition")
BATCH_ALIAS_PAIRS = (
    ("Название", "name"), ("Наименование", "name"), ("База", "database"),
    ("Основа", "balance"), ("Единицы", "units"),
    ("Температура, °C", "temperature_C"), ("Температура", "temperature_C"),
    ("Добавки", "composition"), ("Состав", "composition"),
    ("Давление, Па", "pressure_Pa"), ("Режим стали", "steel_mode"),
    ("Фазы", "phases"),
)
TEMPLATE_HEADERS = (
    "Название", "База", "Основа", "Единицы", "Температура, °C", "Добавки",
    "Режим стали", "Давление, Па", "Фазы",
)
TEMPLATE_XLSX_SHEETS = ("Составы",)
RESULT_XLSX_SHEETS = (
    "Сводка", "Фазовые доли", "Составы фаз ат", "Составы фаз мас", "Исходные данные",
)
CONTENT_KINDS = (
    "alloy-library-json", "project-json", "history-csv", "batch-input-csv",
    "batch-input-xlsx", "batch-template-csv", "batch-template-xlsx",
    "batch-result-xlsx",
)

TICKET_FIELDS = (
    "logical_ref", "sha256", "size_bytes", "binding_digest",
    "binding_generation", "request_digest", "receipt_digest",
    "source_envelope_digest", "content_kind", "content_version",
    "ticket_digest",
)

TEMPLATE_ROWS = (
    ("Учебный Ni–Al", "ni", "NI", "ат.%", 700, "AL=15", "", 101325, ""),
    ("Учебный Al–Cu", "al", "AL", "ат.%", 500, "CU=4", "", 101325, ""),
)

_SHA256_RE = re.compile(r"[0-9a-f]{64}\Z")
_ELEMENT_RE = re.compile(r"[A-Z][A-Z0-9]{0,2}\Z")
_C15 = "C15_LAVES"
_FE_TDB_SHA256 = "236ec4d9b0540de04e4e6305faa208672f31fbdf45b2ae84e92f80bd98053612"
_FE_PROFILE = "thermogar_patch"
_KIND_META = {
    "alloy-library-json": ("alloy-library-json-v1", ".json", "ThermoGar_alloys.json", "application/json", "both"),
    "project-json": ("project-json-v1", ".json", "ThermoGar_project.thermogar.json", "application/json", "both"),
    "history-csv": ("history-csv-v1", ".csv", "ThermoGar_history.csv", "text/csv", "egress"),
    "batch-input-csv": ("batch-input-csv-v1", ".csv", "ThermoGar_batch_input.csv", "text/csv", "ingress"),
    "batch-input-xlsx": ("batch-input-xlsx-v1", ".xlsx", "ThermoGar_batch_input.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "ingress"),
    "batch-template-csv": ("batch-template-csv-v1", ".csv", "ThermoGar_batch_template.csv", "text/csv", "egress"),
    "batch-template-xlsx": ("batch-template-xlsx-v1", ".xlsx", "ThermoGar_batch_template.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "egress"),
    "batch-result-xlsx": ("batch-result-xlsx-v1", ".xlsx", "ThermoGar_batch_result.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "egress"),
}
_UPLOAD_RULES = {
    (ALLOY_UPLOAD_KEY, "alloy-library-json"): ALLOY_UI_TYPES,
    (PROJECT_UPLOAD_KEY, "project-json"): PROJECT_UI_TYPES,
    (BATCH_UPLOAD_KEY, "batch-input-csv"): BATCH_UI_TYPES,
}
_ALIAS_MAP = {alias.strip().casefold(): canonical for alias, canonical in BATCH_ALIAS_PAIRS}
_ALIAS_MAP.update({name.casefold(): name for name in BATCH_REQUIRED_HEADERS + BATCH_OPTIONAL_HEADERS})


class _StateFailure(Exception):
    def __init__(self, reason: verified_loaders.ReasonCode, detail: str) -> None:
        super().__init__(detail)
        self.reason = reason
        self.detail = detail


def _utc(clock: Callable[[], object]) -> str:
    value = clock()
    if isinstance(value, datetime):
        return value.astimezone(timezone.utc).isoformat(timespec="microseconds").replace("+00:00", "Z")
    if type(value) is str and value.endswith("Z"):
        return value
    raise TypeError("clock must return a UTC datetime or Z timestamp")


def _system_clock() -> datetime:
    return datetime.now(timezone.utc)


def _canonical_bytes(value: object) -> bytes:
    return verified_loaders.canonical_json_bytes(value)


def _safe_copy(value: object) -> object:
    return json.loads(_canonical_bytes(value).decode("utf-8"))


def _safe_slug(value: str, fallback: str) -> str:
    normalized = re.sub(r"[^0-9A-Za-zА-Яа-яЁё_-]+", "_", value.strip())
    normalized = normalized.strip("_-")
    return normalized[:80] or fallback


def _digest_without(payload: Mapping[str, object], omitted: str) -> str:
    return verified_loaders.canonical_digest({key: value for key, value in payload.items() if key != omitted})


def _pairs_no_duplicates(pairs: list[tuple[str, object]]) -> dict[str, object]:
    result: dict[str, object] = {}
    for key, value in pairs:
        if key in result:
            raise ValueError(f"duplicate JSON key: {key}")
        result[key] = value
    return result


def _reject_constant(value: str) -> object:
    raise ValueError(f"non-finite JSON constant: {value}")


def _load_json(data: bytes) -> object:
    try:
        text = data.decode("utf-8-sig", errors="strict")
        return json.loads(
            text,
            object_pairs_hook=_pairs_no_duplicates,
            parse_constant=_reject_constant,
        )
    except (UnicodeDecodeError, json.JSONDecodeError, ValueError) as error:
        raise _StateFailure(
            verified_loaders.ReasonCode.IMPORT_SCHEMA_REJECTED,
            f"Strict JSON rejected: {type(error).__name__}.",
        ) from error


def _validate_timestamp(value: object) -> None:
    if type(value) is not str or not value:
        raise ValueError("timestamp is absent")
    parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    if parsed.tzinfo is None or parsed.utcoffset() is None:
        raise ValueError("timestamp lacks timezone")


def _validate_envelope(value: object, kind: str) -> object:
    if type(value) is not dict or set(value) != set(ENVELOPE_KEYS):
        raise _StateFailure(
            verified_loaders.ReasonCode.IMPORT_SCHEMA_REJECTED,
            "Transfer envelope fields are not exact.",
        )
    try:
        if value["schema_version"] != 1 or type(value["schema_version"]) is not int:
            raise ValueError("schema")
        if value["kind"] != kind:
            raise ValueError("kind")
        _validate_timestamp(value["exported_at"])
        expected = value["sha256"]
        if type(expected) is not str or _SHA256_RE.fullmatch(expected) is None:
            raise ValueError("sha256")
        actual = verified_loaders.canonical_digest({key: value[key] for key in ENVELOPE_KEYS if key != "sha256"})
        if actual != expected:
            raise ValueError("digest")
        _canonical_bytes(value["payload"])
    except (TypeError, ValueError, verified_loaders.VerifiedLoaderError) as error:
        raise _StateFailure(
            verified_loaders.ReasonCode.IMPORT_SCHEMA_REJECTED,
            f"Transfer envelope rejected: {error}.",
        ) from error
    return _safe_copy(value["payload"])


def _make_envelope(kind: str, payload: object, clock: Callable[[], object]) -> dict[str, object]:
    safe_payload = _safe_copy(payload)
    envelope: dict[str, object] = {
        "schema_version": 1,
        "kind": kind,
        "exported_at": _utc(clock),
        "payload": safe_payload,
        "sha256": "",
    }
    envelope["sha256"] = verified_loaders.canonical_digest(
        {key: envelope[key] for key in ENVELOPE_KEYS if key != "sha256"}
    )
    return envelope


def _valid_alloy_id(value: object) -> bool:
    if type(value) is str:
        return bool(value.strip())
    if type(value) is bool:
        return value is True
    if type(value) is int:
        return value != 0
    if type(value) is float:
        return math.isfinite(value) and value != 0.0
    return False


def _validate_alloy_payload(value: object) -> list[dict[str, object]]:
    if type(value) is not list:
        raise _StateFailure(verified_loaders.ReasonCode.IMPORT_SCHEMA_REJECTED, "Alloy payload must be a list.")
    result: list[dict[str, object]] = []
    for row in value:
        if type(row) is not dict or not _valid_alloy_id(row.get("id")):
            raise _StateFailure(
                verified_loaders.ReasonCode.IMPORT_SCHEMA_REJECTED,
                "Every alloy row requires a nonempty accepted scalar id.",
            )
        try:
            result.append(_safe_copy(row))
        except verified_loaders.VerifiedLoaderError as error:
            raise _StateFailure(verified_loaders.ReasonCode.IMPORT_SCHEMA_REJECTED, error.detail) from error
    return result


def _validate_project_payload(value: object) -> dict[str, object]:
    if type(value) is not dict or set(value) != set(PROJECT_PAYLOAD_KEYS):
        raise _StateFailure(
            verified_loaders.ReasonCode.IMPORT_SCHEMA_REJECTED,
            "Project payload fields are not the exact 14-key grammar.",
        )
    try:
        expected = {
            "schema_version": 1,
            "kind": "thermogar_project_payload",
            "app_stage": APP_STAGE,
            "app_version": APP_VERSION,
            "release_class": RELEASE_CLASS,
            "software_release_status": SOFTWARE_RELEASE_STATUS,
            "scientific_material_status": SCIENTIFIC_MATERIAL_STATUS,
            "production_use": PRODUCTION_USE,
        }
        if any(value[key] != expected_value for key, expected_value in expected.items()):
            raise ValueError("project identity")
        if type(value["name"]) is not str or not value["name"].strip() or len(value["name"].strip()) > 200:
            raise ValueError("project name")
        if type(value["description"]) is not str or value["widget_state"] != {}:
            raise ValueError("project body")
        _validate_timestamp(value["created_at"])
        _validate_timestamp(value["updated_at"])
        context = value["context"]
        if type(context) is not dict:
            raise ValueError("project context")
        for item in context.values():
            if type(item) is float and not math.isfinite(item):
                raise ValueError("nonfinite context")
        database_key = context.get("database_key")
        if database_key not in ("ni", "al", "fe"):
            raise ValueError("database key")
        if database_key == "fe":
            if context.get("fe_profile_key") != _FE_PROFILE:
                raise ValueError("Fe profile")
            if context.get("database_sha256") != _FE_TDB_SHA256:
                raise ValueError("Fe digest")
        elif context.get("fe_profile_key") is not None:
            raise ValueError("non-Fe profile")
        safe = _safe_copy(value)
        safe["name"] = value["name"].strip()
        safe["description"] = value["description"].strip()
        return safe
    except (KeyError, TypeError, ValueError, verified_loaders.VerifiedLoaderError) as error:
        raise _StateFailure(
            verified_loaders.ReasonCode.IMPORT_SCHEMA_REJECTED,
            f"Project payload rejected: {error}.",
        ) from error


def _canonical_header(value: object) -> str:
    text = str(value).strip()
    return _ALIAS_MAP.get(text.casefold(), text.upper() if _ELEMENT_RE.fullmatch(text.upper()) else text)


def _number(value: object, label: str, *, positive: bool = False) -> float:
    if type(value) is bool:
        raise ValueError(label)
    result = float(value)
    if not math.isfinite(result) or (positive and result <= 0.0):
        raise ValueError(label)
    return result


def _units(value: object) -> str:
    text = str(value).strip().casefold().replace(" ", "").replace(".", "")
    if text in {"at", "at%", "ат%", "атомные%", "атомные", "atomic"}:
        return "at"
    if text in {"wt", "wt%", "мас%", "массовые%", "массовые", "mass"}:
        return "wt"
    raise ValueError("units")


def _steel_mode(value: object) -> str:
    text = str(value or "").strip().casefold()
    if any(token in text for token in ("стаб", "граф", "stable", "graphite")):
        return "stable"
    return "metastable"


def _phase_tokens(value: object) -> tuple[str, ...]:
    if value is None or str(value).strip() == "":
        return ()
    tokens = tuple(item.strip().upper() for item in re.split(r"[,;]", str(value)) if item.strip())
    if any(not token or "\\" in token or "/" in token for token in tokens):
        raise ValueError("phase tokens")
    if _C15 in tokens:
        raise _StateFailure(
            verified_loaders.ReasonCode.C15_PHASE_REJECTED,
            "C15_LAVES is rejected before state persistence.",
        )
    return tokens


def _normalize_batch_table(headers: Sequence[object], source_rows: Sequence[Sequence[object]]) -> dict[str, object]:
    canonical_headers = tuple(_canonical_header(item) for item in headers)
    if len(canonical_headers) < 5 or len(canonical_headers) > 256 or len(set(canonical_headers)) != len(canonical_headers):
        raise ValueError("batch headers")
    if not set(BATCH_REQUIRED_HEADERS).issubset(canonical_headers):
        raise ValueError("required headers")
    element_headers = tuple(sorted(header for header in canonical_headers if _ELEMENT_RE.fullmatch(header)))
    unknown = set(canonical_headers) - set(BATCH_REQUIRED_HEADERS) - set(BATCH_OPTIONAL_HEADERS) - set(element_headers)
    if unknown:
        raise ValueError("unknown headers")
    columns = BATCH_REQUIRED_HEADERS + BATCH_OPTIONAL_HEADERS + element_headers
    if not source_rows or len(source_rows) > 100:
        raise ValueError("batch row count")
    normalized_rows: list[list[object]] = []
    for values in source_rows:
        padded = list(values) + [None] * (len(canonical_headers) - len(values))
        row = dict(zip(canonical_headers, padded[:len(canonical_headers)]))
        name = str(row.get("name", "")).strip()
        database = str(row.get("database", "")).strip().casefold()
        balance = str(row.get("balance", "")).strip().upper()
        if not name or database not in ("ni", "al", "fe") or _ELEMENT_RE.fullmatch(balance) is None:
            raise ValueError("batch identity")
        temperature = _number(row.get("temperature_C"), "temperature")
        pressure_value = row.get("pressure_Pa")
        pressure = 101325.0 if pressure_value in (None, "") else _number(pressure_value, "pressure", positive=True)
        phases = ",".join(_phase_tokens(row.get("phases")))
        composition = "" if row.get("composition") is None else str(row.get("composition")).strip()
        clean: dict[str, object] = {
            "name": name,
            "database": database,
            "balance": balance,
            "units": _units(row.get("units")),
            "temperature_C": temperature,
            "pressure_Pa": pressure,
            "steel_mode": _steel_mode(row.get("steel_mode")),
            "phases": phases,
            "composition": composition,
        }
        for element in element_headers:
            value = row.get(element)
            clean[element] = "" if value in (None, "") else _number(value, element)
        normalized_rows.append([clean[column] for column in columns])
    return {"columns": list(columns), "rows": normalized_rows}


def _parse_csv(data: bytes) -> dict[str, object]:
    try:
        text = data.decode("utf-8-sig", errors="strict")
    except UnicodeDecodeError as error:
        raise _StateFailure(verified_loaders.ReasonCode.IMPORT_SCHEMA_REJECTED, "Batch CSV must be strict UTF-8/SIG.") from error
    valid: list[dict[str, object]] = []
    c15_rejected = False
    for separator in (",", ";"):
        try:
            rows = list(csv.reader(StringIO(text, newline=""), delimiter=separator, strict=True))
            if not rows:
                raise ValueError("empty")
            valid.append(_normalize_batch_table(rows[0], rows[1:]))
        except _StateFailure as error:
            if error.reason is verified_loaders.ReasonCode.C15_PHASE_REJECTED:
                c15_rejected = True
        except (csv.Error, TypeError, ValueError):
            continue
    if c15_rejected:
        raise _StateFailure(
            verified_loaders.ReasonCode.C15_PHASE_REJECTED,
            "C15_LAVES is rejected before state persistence.",
        )
    if len(valid) != 1:
        raise _StateFailure(
            verified_loaders.ReasonCode.IMPORT_SCHEMA_REJECTED,
            "Exactly one comma/semicolon CSV parse must satisfy the batch grammar.",
        )
    return valid[0]


def _xlsx_has_forbidden_members(data: bytes) -> bool:
    try:
        with ZipFile(BytesIO(data), "r") as archive:
            names = tuple(name.casefold() for name in archive.namelist())
    except BadZipFile as error:
        raise _StateFailure(verified_loaders.ReasonCode.IMPORT_SCHEMA_REJECTED, "XLSX is not a valid OOXML archive.") from error
    return any("vbaproject" in name or "/externallinks/" in name for name in names)


def _parse_xlsx(data: bytes) -> dict[str, object]:
    if _xlsx_has_forbidden_members(data):
        raise _StateFailure(verified_loaders.ReasonCode.IMPORT_SCHEMA_REJECTED, "XLSX macros/external links are rejected.")
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(data), read_only=False, data_only=False, keep_vba=False, keep_links=False)
        if len(workbook.worksheets) != 1 or workbook.worksheets[0].sheet_state != "visible":
            raise ValueError("worksheet count/state")
        if getattr(workbook, "_external_links", ()):
            raise ValueError("external links")
        sheet = workbook.worksheets[0]
        if not sheet.title.strip():
            raise ValueError("sheet title")
        rows = list(sheet.iter_rows(values_only=False))
        if not rows:
            raise ValueError("empty workbook")
        for row in rows:
            for cell in row:
                if cell.data_type == "f" or (type(cell.value) is str and cell.value.startswith("=")):
                    raise ValueError("formula")
        values = [[cell.value for cell in row] for row in rows]
        return _normalize_batch_table(values[0], values[1:])
    except _StateFailure:
        raise
    except Exception as error:
        raise _StateFailure(
            verified_loaders.ReasonCode.IMPORT_SCHEMA_REJECTED,
            f"Safe XLSX grammar rejected: {type(error).__name__}.",
        ) from error
    finally:
        if "workbook" in locals():
            workbook.close()


def _batch_csv_bytes(table: Mapping[str, object]) -> bytes:
    columns = tuple(table["columns"])
    rows = tuple(tuple(row) for row in table["rows"])
    output = StringIO(newline="")
    writer = csv.writer(output, delimiter=",", lineterminator="\n")
    writer.writerow(columns)
    writer.writerows(rows)
    return output.getvalue().encode("utf-8")


def _history_csv_bytes(rows: Sequence[Mapping[str, object]]) -> bytes:
    output = StringIO(newline="")
    writer = csv.writer(output, delimiter=",", lineterminator="\n")
    writer.writerow(HISTORY_HEADERS)
    for row in rows:
        if type(row) is not dict or set(row) != set(HISTORY_HEADERS):
            raise ValueError("history headers")
        writer.writerow([str(row[header]) for header in HISTORY_HEADERS])
    return b"\xef\xbb\xbf" + output.getvalue().encode("utf-8")


def _table_value(headers: Sequence[str], rows: Sequence[Sequence[object]]) -> dict[str, object]:
    value = {"columns": list(headers), "rows": [list(row) for row in rows]}
    return _safe_copy(value)


def batch_template_value() -> dict[str, object]:
    return _table_value(TEMPLATE_HEADERS, TEMPLATE_ROWS)


def _validate_sheet_value(value: object, *, max_rows: int) -> dict[str, object]:
    if type(value) is not dict or tuple(value) != ("columns", "rows"):
        raise ValueError("sheet mapping")
    columns = value["columns"]
    rows = value["rows"]
    if type(columns) is not list or not columns or len(columns) > 256 or len(set(columns)) != len(columns):
        raise ValueError("sheet columns")
    if any(type(column) is not str or not column for column in columns):
        raise ValueError("sheet column labels")
    if type(rows) is not list or len(rows) > max_rows:
        raise ValueError("sheet rows")
    for row in rows:
        if type(row) is not list or len(row) != len(columns):
            raise ValueError("sheet row width")
        _canonical_bytes(row)
    return _safe_copy(value)


def batch_result_value(sheets: Mapping[str, object]) -> dict[str, object]:
    if type(sheets) is not dict or tuple(sheets) != RESULT_XLSX_SHEETS:
        raise ValueError("result sheet sequence")
    result: dict[str, object] = {}
    for name in RESULT_XLSX_SHEETS:
        maximum = 100 if name in ("Сводка", "Исходные данные") else 13200
        result[name] = _validate_sheet_value(sheets[name], max_rows=maximum)
    return result


def _xlsx_bytes(sheets: Mapping[str, Mapping[str, object]]) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    first = True
    for name, table in sheets.items():
        sheet = workbook.active if first else workbook.create_sheet()
        first = False
        sheet.title = name
        sheet.append(list(table["columns"]))
        for row in table["rows"]:
            sheet.append(list(row))
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def semantic_digest_for(content_kind: str, value: object) -> str:
    if content_kind == "alloy-library-json":
        semantic = _validate_alloy_payload(value)
        return verified_loaders.canonical_digest(semantic)
    if content_kind == "project-json":
        semantic = _validate_project_payload(value)
        return verified_loaders.canonical_digest(semantic)
    if content_kind == "history-csv":
        if not isinstance(value, Sequence) or isinstance(value, (str, bytes)):
            raise ValueError("history rows")
        return hashlib.sha256(_history_csv_bytes(value)).hexdigest()
    if content_kind in ("batch-input-csv", "batch-input-xlsx"):
        if type(value) is not dict:
            raise ValueError("batch table")
        return hashlib.sha256(_batch_csv_bytes(value)).hexdigest()
    if content_kind in ("batch-template-csv", "batch-template-xlsx"):
        table = _validate_sheet_value(value, max_rows=2)
        if table != batch_template_value():
            raise ValueError("template table")
        return hashlib.sha256(_batch_csv_bytes(table)).hexdigest()
    if content_kind == "batch-result-xlsx":
        return verified_loaders.canonical_digest(batch_result_value(value))
    raise ValueError("unknown content kind")


def _deep_has_c15(value: object) -> bool:
    if type(value) is str:
        return value.strip().upper() == _C15
    if type(value) is list:
        return any(_deep_has_c15(item) for item in value)
    if type(value) is dict:
        return any(_deep_has_c15(item) for item in value.values())
    return False


def _derive_request(
    request: verified_loaders.FeatureRequest,
    content_kind: str,
    semantic_digest: str,
) -> verified_loaders.FeatureRequest:
    inputs = dict(request.inputs)
    existing = inputs.get("semantic_digest")
    if existing is not None and existing != semantic_digest:
        raise _StateFailure(verified_loaders.ReasonCode.REQUEST_DIGEST_MISMATCH, "Prepared semantic digest mismatch.")
    inputs["content_kind"] = content_kind
    inputs["semantic_digest"] = semantic_digest
    inputs = _safe_copy(inputs)
    provisional = {
        "schema": request.schema,
        "feature_id": request.feature_id,
        "feature_revision": request.feature_revision,
        "binding_digest": request.binding_digest,
        "binding_generation": request.binding_generation,
        "inputs": inputs,
        "inputs_digest": verified_loaders.canonical_digest(inputs),
        "requested_phases": list(request.requested_phases),
        "requested_phases_digest": request.requested_phases_digest,
        "effective_phases": list(request.effective_phases),
        "effective_phases_digest": request.effective_phases_digest,
        "request_digest": "",
    }
    provisional["request_digest"] = _digest_without(provisional, "request_digest")
    return verified_loaders.FeatureRequest(
        schema=provisional["schema"],
        feature_id=provisional["feature_id"],
        feature_revision=provisional["feature_revision"],
        binding_digest=provisional["binding_digest"],
        binding_generation=provisional["binding_generation"],
        inputs=provisional["inputs"],
        inputs_digest=provisional["inputs_digest"],
        requested_phases=request.requested_phases,
        requested_phases_digest=provisional["requested_phases_digest"],
        effective_phases=request.effective_phases,
        effective_phases_digest=provisional["effective_phases_digest"],
        request_digest=provisional["request_digest"],
    )


def _make_rejection(
    request: verified_loaders.FeatureRequest,
    reason: verified_loaders.ReasonCode,
    detail: str,
    clock: Callable[[], object],
) -> verified_loaders.RejectedFeatureReceipt:
    payload = {
        "schema": verified_loaders.SCHEMA_REJECTION,
        "feature_id": request.feature_id,
        "feature_revision": request.feature_revision,
        "outcome": "rejected",
        "reason_code": reason.value,
        "reason_detail": detail[:verified_loaders.MAX_REASON_DETAIL_CHARS],
        "binding_digest": request.binding_digest,
        "binding_generation": request.binding_generation,
        "inputs_digest": request.inputs_digest,
        "requested_phases_digest": request.requested_phases_digest,
        "effective_phases_digest": request.effective_phases_digest,
        "request_digest": request.request_digest,
        "backend_calls": 0,
        "rejected_at_utc": _utc(clock),
        "receipt_digest": "",
    }
    payload["receipt_digest"] = _digest_without(payload, "receipt_digest")
    return verified_loaders.RejectedFeatureReceipt(**payload)


@dataclass(frozen=True, slots=True)
class VerifiedArtifactRef:
    logical_ref: str
    sha256: str
    size_bytes: int
    binding_digest: str
    binding_generation: int
    request_digest: str
    receipt_digest: str
    source_envelope_digest: str
    content_kind: str
    content_version: str
    ticket_digest: str

    def __post_init__(self) -> None:
        if tuple(field.name for field in fields(self)) != TICKET_FIELDS:
            raise ValueError("ticket field order")
        if self.content_kind not in CONTENT_KINDS or self.content_version != "1":
            raise ValueError("ticket kind/version")
        for value in (
            self.sha256, self.binding_digest, self.request_digest,
            self.receipt_digest, self.source_envelope_digest, self.ticket_digest,
        ):
            if type(value) is not str or _SHA256_RE.fullmatch(value) is None:
                raise ValueError("ticket digest")
        if type(self.binding_generation) is not int or self.binding_generation <= 0:
            raise ValueError("ticket generation")
        if type(self.size_bytes) is not int or self.size_bytes < 0 or self.size_bytes > MAX_WORKSPACE_FILE_BYTES:
            raise ValueError("ticket size")
        parent, extension, _name, _media, _direction = _KIND_META[self.content_kind]
        expected_ref = f"state/{parent}/{self.sha256}{extension}"
        if self.logical_ref != expected_ref or "\\" in self.logical_ref or ":" in self.logical_ref:
            raise ValueError("ticket logical reference")
        if self.ticket_digest != _digest_without(self.to_dict(), "ticket_digest"):
            raise ValueError("ticket digest mismatch")

    def to_dict(self) -> dict[str, object]:
        return {field: getattr(self, field) for field in TICKET_FIELDS}


@dataclass(frozen=True, slots=True)
class _StateRecord:
    request: verified_loaders.FeatureRequest
    value: object
    ticket: VerifiedArtifactRef
    artifact_entry: dict[str, object]
    source_envelope: verified_loaders.ResultEnvelope
    result_envelope: verified_loaders.ResultEnvelope
    destination: Path


def _result_envelope(
    request: verified_loaders.FeatureRequest,
    receipt_digest: str,
    settings: dict[str, object],
    artifacts: Sequence[Mapping[str, object]],
    clock: Callable[[], object],
) -> verified_loaders.ResultEnvelope:
    safe_settings = _safe_copy(settings)
    safe_artifacts = [_safe_copy(dict(item)) for item in artifacts]
    settings_digest = verified_loaders.canonical_digest(safe_settings)
    tables_digest = verified_loaders.canonical_digest([])
    figures_digest = verified_loaders.canonical_digest([])
    artifacts_digest = verified_loaders.canonical_digest(safe_artifacts)
    result_digest = verified_loaders.canonical_digest(
        {
            "settings_digest": settings_digest,
            "tables_digest": tables_digest,
            "figures_digest": figures_digest,
            "artifacts_digest": artifacts_digest,
        }
    )
    payload = {
        "schema": verified_loaders.SCHEMA_RESULT_ENVELOPE,
        "feature_id": request.feature_id,
        "feature_revision": request.feature_revision,
        "binding_digest": request.binding_digest,
        "binding_generation": request.binding_generation,
        "request_digest": request.request_digest,
        "receipt_digest": receipt_digest,
        "outcome": "success",
        "settings": safe_settings,
        "settings_digest": settings_digest,
        "tables": (),
        "tables_digest": tables_digest,
        "figures": (),
        "figures_digest": figures_digest,
        "artifacts": tuple(safe_artifacts),
        "artifacts_digest": artifacts_digest,
        "result_digest": result_digest,
        "created_at_utc": _utc(clock),
        "envelope_digest": "",
    }
    payload["envelope_digest"] = _digest_without(
        {
            **payload,
            "tables": [],
            "figures": [],
            "artifacts": safe_artifacts,
        },
        "envelope_digest",
    )
    return verified_loaders.ResultEnvelope(**payload)


class StateStore:
    """Finite state owner. Upload/download bodies and paths never escape it."""

    def __init__(
        self,
        paths: object,
        ui: object,
        *,
        binding_probe: Callable[[], tuple[str, int]] | None = None,
        clock: Callable[[], object] = _system_clock,
    ) -> None:
        state_root = getattr(paths, "state_root", None)
        if not isinstance(state_root, Path) or not state_root.is_absolute():
            raise TypeError("StateStore requires injected resolved paths.state_root.")
        if not callable(getattr(ui, "file_uploader", None)) or not callable(getattr(ui, "download_button", None)):
            raise TypeError("StateStore requires a narrow uploader/download UI.")
        self._paths = paths
        self._ui = ui
        self._binding_probe = binding_probe
        self._clock = clock
        self._records: dict[str, _StateRecord] = {}
        self._counters = {
            "uploader_calls": 0,
            "parse_calls": 0,
            "policy_calls": 0,
            "write_calls": 0,
            "ticket_calls": 0,
            "backend_calls": 0,
        }

    @property
    def counters(self) -> dict[str, int]:
        return dict(self._counters)

    def _live(self, binding_digest: str, generation: int) -> None:
        if self._binding_probe is None:
            return
        current_digest, current_generation = self._binding_probe()
        if current_digest != binding_digest:
            raise _StateFailure(verified_loaders.ReasonCode.BINDING_STALE, "State binding is stale.")
        if current_generation != generation:
            raise _StateFailure(verified_loaders.ReasonCode.GENERATION_STALE, "State generation is stale.")

    def _parse_ingress(self, kind: str, data: bytes, upload_name: str) -> tuple[str, object, str]:
        self._counters["parse_calls"] += 1
        actual_kind = kind
        if kind == "batch-input-csv":
            suffix = Path(upload_name).suffix.casefold()
            if suffix == ".xlsx":
                actual_kind = "batch-input-xlsx"
            elif suffix != ".csv":
                raise _StateFailure(verified_loaders.ReasonCode.IMPORT_SCHEMA_REJECTED, "Batch upload must be .csv or .xlsx.")
        if actual_kind == "alloy-library-json":
            value = _validate_alloy_payload(_validate_envelope(_load_json(data), "thermogar_alloys"))
        elif actual_kind == "project-json":
            value = _validate_project_payload(_validate_envelope(_load_json(data), "thermogar_project"))
        elif actual_kind == "batch-input-csv":
            value = _parse_csv(data)
        elif actual_kind == "batch-input-xlsx":
            value = _parse_xlsx(data)
        else:
            raise _StateFailure(verified_loaders.ReasonCode.CAPABILITY_UNAVAILABLE, "Content kind is egress-only.")
        self._counters["policy_calls"] += 1
        if _deep_has_c15(value):
            raise _StateFailure(verified_loaders.ReasonCode.C15_PHASE_REJECTED, "C15_LAVES is rejected before state persistence.")
        semantic_digest = semantic_digest_for(actual_kind, value)
        return actual_kind, value, semantic_digest

    def _source_envelope(
        self,
        request: verified_loaders.FeatureRequest,
        content_kind: str,
        semantic_digest: str,
    ) -> tuple[str, verified_loaders.ResultEnvelope]:
        receipt_digest = verified_loaders.canonical_digest(
            {
                "binding_digest": request.binding_digest,
                "binding_generation": request.binding_generation,
                "content_kind": content_kind,
                "request_digest": request.request_digest,
                "semantic_digest": semantic_digest,
                "state_role": "verified-source",
            }
        )
        envelope = _result_envelope(
            request,
            receipt_digest,
            {"content_kind": content_kind, "semantic_digest": semantic_digest},
            (),
            self._clock,
        )
        return receipt_digest, envelope

    def _persist(self, content_kind: str, data: bytes) -> tuple[Path, str, int]:
        raw_sha256 = hashlib.sha256(data).hexdigest()
        raw_size = len(data)
        parent_name, extension, _name, _media, _direction = _KIND_META[content_kind]
        state_root = self._paths.state_root
        finite_parent = state_root / "state" / parent_name
        destination = finite_parent / f"{raw_sha256}{extension}"
        ensure_plain_directory(state_root)
        ensure_plain_directory(finite_parent)
        self._counters["write_calls"] += 1
        try:
            atomic_write_bytes(
                destination,
                data,
                create_backup=False,
                overwrite=False,
                canonical_root=state_root,
            )
        except FileExistsError:
            try:
                snapshot = read_verified_snapshot(
                    destination,
                    expected_sha256=raw_sha256,
                    maximum_bytes=MAX_WORKSPACE_FILE_BYTES,
                    canonical_root=state_root,
                )
            except Exception as error:
                raise _StateFailure(verified_loaders.ReasonCode.STATE_CONFLICT, f"Existing state artifact conflict: {type(error).__name__}.") from error
            if snapshot.sha256 != raw_sha256 or snapshot.size != raw_size:
                raise _StateFailure(verified_loaders.ReasonCode.STATE_CONFLICT, "Existing state artifact SHA/size mismatch.")
        except Exception as error:
            raise _StateFailure(verified_loaders.ReasonCode.ARTIFACT_WRITE_FAILED, f"State create-only write failed: {type(error).__name__}.") from error
        try:
            snapshot = read_verified_snapshot(
                destination,
                expected_sha256=raw_sha256,
                maximum_bytes=MAX_WORKSPACE_FILE_BYTES,
                canonical_root=state_root,
            )
        except Exception as error:
            raise _StateFailure(verified_loaders.ReasonCode.ARTIFACT_WRITE_FAILED, f"State post-write snapshot failed: {type(error).__name__}.") from error
        if snapshot.sha256 != raw_sha256 or snapshot.size != raw_size:
            raise _StateFailure(verified_loaders.ReasonCode.ARTIFACT_WRITE_FAILED, "State post-write SHA/size mismatch.")
        return destination, raw_sha256, raw_size

    def _mint(
        self,
        request: verified_loaders.FeatureRequest,
        content_kind: str,
        value: object,
        destination: Path,
        raw_sha256: str,
        raw_size: int,
        receipt_digest: str,
        source_envelope: verified_loaders.ResultEnvelope,
    ) -> VerifiedArtifactRef:
        self._counters["ticket_calls"] += 1
        parent_name, extension, file_name, media_type, _direction = _KIND_META[content_kind]
        if content_kind == "project-json":
            file_name = (
                _safe_slug(str(value["name"]), "ThermoGar_project")
                + ".thermogar.json"
            )
        logical_ref = f"state/{parent_name}/{raw_sha256}{extension}"
        provisional = {
            "logical_ref": logical_ref,
            "sha256": raw_sha256,
            "size_bytes": raw_size,
            "binding_digest": request.binding_digest,
            "binding_generation": request.binding_generation,
            "request_digest": request.request_digest,
            "receipt_digest": receipt_digest,
            "source_envelope_digest": source_envelope.envelope_digest,
            "content_kind": content_kind,
            "content_version": "1",
            "ticket_digest": "",
        }
        provisional["ticket_digest"] = _digest_without(provisional, "ticket_digest")
        ticket = VerifiedArtifactRef(**provisional)
        artifact_entry = {
            "name": file_name,
            "media_type": media_type,
            "sha256": raw_sha256,
            "size_bytes": raw_size,
            "payload_ref": logical_ref,
        }
        result_envelope = _result_envelope(
            request,
            receipt_digest,
            {"transfer_ticket": ticket.to_dict()},
            (artifact_entry,),
            self._clock,
        )
        self._records[ticket.ticket_digest] = _StateRecord(
            request=request,
            value=_safe_copy(value),
            ticket=ticket,
            artifact_entry=artifact_entry,
            source_envelope=source_envelope,
            result_envelope=result_envelope,
            destination=destination,
        )
        return ticket

    def ingest_from_widget(
        self,
        request: verified_loaders.FeatureRequest,
        content_kind: str,
        label: str,
        types: tuple[str, ...],
        key: str,
    ) -> VerifiedArtifactRef | verified_loaders.RejectedFeatureReceipt:
        if type(request) is not verified_loaders.FeatureRequest:
            raise TypeError("ingress requires a prepared FeatureRequest")
        if type(content_kind) is not str or (key, content_kind) not in _UPLOAD_RULES:
            return _make_rejection(request, verified_loaders.ReasonCode.IMPORT_SCHEMA_REJECTED, "Uploader kind/key is outside the finite map.", self._clock)
        if types != _UPLOAD_RULES[(key, content_kind)] or type(label) is not str or not label:
            return _make_rejection(request, verified_loaders.ReasonCode.IMPORT_SCHEMA_REJECTED, "Uploader scalar metadata mismatch.", self._clock)
        if _C15 in request.requested_phases or _C15 in request.effective_phases:
            return _make_rejection(request, verified_loaders.ReasonCode.C15_PHASE_REJECTED, "C15_LAVES is rejected before uploader dispatch.", self._clock)
        self._counters["uploader_calls"] += 1
        uploaded = self._ui.file_uploader(label, type=types, key=key)
        if uploaded is None:
            return _make_rejection(request, verified_loaders.ReasonCode.USER_INPUT_REQUIRED, "Select one verified state artifact.", self._clock)
        chunks: list[bytes] = []
        total = 0
        try:
            while True:
                chunk = uploaded.read(65536)
                if chunk in (b"", None):
                    break
                if type(chunk) is not bytes:
                    raise TypeError("upload chunks must be bytes")
                total += len(chunk)
                if total > MAX_WORKSPACE_FILE_BYTES:
                    return _make_rejection(request, verified_loaders.ReasonCode.ARTIFACT_OVERSIZE, "State artifact exceeds 67108864 bytes.", self._clock)
                chunks.append(chunk)
            data = b"".join(chunks)
            actual_kind, value, semantic_digest = self._parse_ingress(
                content_kind,
                data,
                str(getattr(uploaded, "name", "")),
            )
            derived = _derive_request(request, actual_kind, semantic_digest)
            self._live(derived.binding_digest, derived.binding_generation)
            receipt_digest, source_envelope = self._source_envelope(derived, actual_kind, semantic_digest)
            destination, raw_sha256, raw_size = self._persist(actual_kind, data)
            return self._mint(
                derived, actual_kind, value, destination, raw_sha256, raw_size,
                receipt_digest, source_envelope,
            )
        except _StateFailure as error:
            return _make_rejection(request, error.reason, error.detail, self._clock)
        except Exception as error:
            return _make_rejection(request, verified_loaders.ReasonCode.IMPORT_SCHEMA_REJECTED, f"State ingress failed: {type(error).__name__}.", self._clock)

    def prepare_egress(
        self,
        request: verified_loaders.FeatureRequest,
        content_kind: str,
        value: object,
    ) -> VerifiedArtifactRef | verified_loaders.RejectedFeatureReceipt:
        if type(request) is not verified_loaders.FeatureRequest:
            raise TypeError("egress requires a prepared FeatureRequest")
        if content_kind not in CONTENT_KINDS or _KIND_META[content_kind][4] not in ("egress", "both"):
            return _make_rejection(request, verified_loaders.ReasonCode.CAPABILITY_UNAVAILABLE, "Content kind is ingress-only.", self._clock)
        try:
            self._counters["policy_calls"] += 1
            if _C15 in request.requested_phases or _C15 in request.effective_phases or _deep_has_c15(value):
                raise _StateFailure(verified_loaders.ReasonCode.C15_PHASE_REJECTED, "C15_LAVES is rejected before state persistence.")
            if content_kind == "alloy-library-json":
                safe_value = _validate_alloy_payload(value)
                data = _canonical_bytes(_make_envelope("thermogar_alloys", safe_value, self._clock))
            elif content_kind == "project-json":
                safe_value = _validate_project_payload(value)
                data = _canonical_bytes(_make_envelope("thermogar_project", safe_value, self._clock))
            elif content_kind == "history-csv":
                safe_value = _safe_copy(value)
                data = _history_csv_bytes(safe_value)
            elif content_kind == "batch-template-csv":
                safe_value = _validate_sheet_value(value, max_rows=2)
                if safe_value != batch_template_value():
                    raise ValueError("template rows")
                data = b"\xef\xbb\xbf" + _batch_csv_bytes(safe_value)
            elif content_kind == "batch-template-xlsx":
                safe_value = _validate_sheet_value(value, max_rows=2)
                if safe_value != batch_template_value():
                    raise ValueError("template rows")
                data = _xlsx_bytes({TEMPLATE_XLSX_SHEETS[0]: safe_value})
            elif content_kind == "batch-result-xlsx":
                safe_value = batch_result_value(value)
                data = _xlsx_bytes(safe_value)
            else:
                raise _StateFailure(verified_loaders.ReasonCode.CAPABILITY_UNAVAILABLE, "Content kind is ingress-only.")
            semantic_digest = semantic_digest_for(content_kind, safe_value)
            derived = _derive_request(request, content_kind, semantic_digest)
            self._live(derived.binding_digest, derived.binding_generation)
            receipt_digest, source_envelope = self._source_envelope(derived, content_kind, semantic_digest)
            destination, raw_sha256, raw_size = self._persist(content_kind, data)
            return self._mint(
                derived, content_kind, safe_value, destination, raw_sha256,
                raw_size, receipt_digest, source_envelope,
            )
        except _StateFailure as error:
            return _make_rejection(request, error.reason, error.detail, self._clock)
        except Exception as error:
            return _make_rejection(request, verified_loaders.ReasonCode.EXPORT_SOURCE_REJECTED, f"State egress rejected: {type(error).__name__}.", self._clock)

    def _record(
        self,
        ticket: VerifiedArtifactRef,
        source_envelope_digest: str,
    ) -> _StateRecord:
        if type(ticket) is not VerifiedArtifactRef:
            raise verified_loaders.VerifiedLoaderError(verified_loaders.ReasonCode.RAW_PATH_REJECTED, "State access requires a private ticket.")
        record = self._records.get(ticket.ticket_digest)
        if record is None or record.ticket != ticket:
            raise verified_loaders.VerifiedLoaderError(verified_loaders.ReasonCode.ARTIFACT_MISSING, "State ticket is unknown in this process.")
        if source_envelope_digest != ticket.source_envelope_digest or source_envelope_digest != record.source_envelope.envelope_digest:
            raise verified_loaders.VerifiedLoaderError(verified_loaders.ReasonCode.ENVELOPE_CONTEXT_MISMATCH, "Source envelope digest mismatch.")
        try:
            self._live(ticket.binding_digest, ticket.binding_generation)
        except _StateFailure as error:
            raise verified_loaders.VerifiedLoaderError(error.reason, error.detail) from error
        if record.request.request_digest != ticket.request_digest or record.source_envelope.receipt_digest != ticket.receipt_digest:
            raise verified_loaders.VerifiedLoaderError(verified_loaders.ReasonCode.ENVELOPE_INVALID, "Ticket request/receipt mismatch.")
        return record

    def paired_artifact(
        self,
        ticket: VerifiedArtifactRef,
    ) -> tuple[dict[str, object], verified_loaders.ResultEnvelope]:
        record = self._records.get(ticket.ticket_digest)
        if record is None or record.ticket != ticket:
            raise verified_loaders.VerifiedLoaderError(verified_loaders.ReasonCode.ARTIFACT_MISSING, "State ticket is unknown.")
        return dict(record.artifact_entry), record.result_envelope

    def canonical_value(
        self,
        ticket: VerifiedArtifactRef,
        source_envelope_digest: str,
    ) -> object:
        record = self._record(ticket, source_envelope_digest)
        try:
            snapshot = read_verified_snapshot(
                record.destination,
                expected_sha256=ticket.sha256,
                maximum_bytes=MAX_WORKSPACE_FILE_BYTES,
                canonical_root=self._paths.state_root,
            )
        except Exception as error:
            raise verified_loaders.VerifiedLoaderError(verified_loaders.ReasonCode.ARTIFACT_IO_FAILED, f"State snapshot failed: {type(error).__name__}.") from error
        if snapshot.sha256 != ticket.sha256 or snapshot.size != ticket.size_bytes:
            raise verified_loaders.VerifiedLoaderError(verified_loaders.ReasonCode.ARTIFACT_IO_FAILED, "State snapshot SHA/size mismatch.")
        return _safe_copy(record.value)

    def restore_context(
        self,
        ticket: VerifiedArtifactRef,
        source_envelope_digest: str,
        rebind: Callable[[dict[str, object]], dict[str, object]],
        *,
        alloy_index: int | None = None,
    ) -> dict[str, object]:
        value = self.canonical_value(ticket, source_envelope_digest)
        if ticket.content_kind == "project-json":
            context = value["context"]
        elif ticket.content_kind == "alloy-library-json" and type(alloy_index) is int:
            context = value[alloy_index]
        else:
            raise verified_loaders.VerifiedLoaderError(verified_loaders.ReasonCode.IMPORT_SCHEMA_REJECTED, "Artifact has no restorable context.")
        if type(context) is not dict or not callable(rebind):
            raise verified_loaders.VerifiedLoaderError(verified_loaders.ReasonCode.BINDING_IDENTITY_MISMATCH, "Canonical rebind is unavailable.")
        rebound = rebind(_safe_copy(context))
        if type(rebound) is not dict:
            raise verified_loaders.VerifiedLoaderError(verified_loaders.ReasonCode.BINDING_IDENTITY_MISMATCH, "Canonical rebind returned invalid evidence.")
        return _safe_copy(rebound)

    def render_download(
        self,
        ticket: VerifiedArtifactRef,
        source_envelope_digest: str,
        label: str,
        *,
        key: str,
    ) -> bool:
        record = self._record(ticket, source_envelope_digest)
        try:
            with held_verified_snapshot(
                record.destination,
                expected_sha256=ticket.sha256,
                maximum_bytes=MAX_WORKSPACE_FILE_BYTES,
                canonical_root=self._paths.state_root,
            ) as snapshot:
                if snapshot.sha256 != ticket.sha256 or snapshot.size != ticket.size_bytes:
                    raise verified_loaders.VerifiedLoaderError(verified_loaders.ReasonCode.ARTIFACT_IO_FAILED, "Download snapshot SHA/size mismatch.")
                return bool(
                    self._ui.download_button(
                        label,
                        data=snapshot.data,
                        file_name=record.artifact_entry["name"],
                        mime=record.artifact_entry["media_type"],
                        key=key,
                    )
                )
        except verified_loaders.VerifiedLoaderError:
            raise
        except Exception as error:
            raise verified_loaders.VerifiedLoaderError(verified_loaders.ReasonCode.ARTIFACT_IO_FAILED, f"Download snapshot failed: {type(error).__name__}.") from error


__all__ = [
    "ALLOY_UI_TYPES", "PROJECT_UI_TYPES", "BATCH_UI_TYPES",
    "ALLOY_UPLOAD_KEY", "PROJECT_UPLOAD_KEY", "BATCH_UPLOAD_KEY",
    "ENVELOPE_KEYS", "PROJECT_PAYLOAD_KEYS", "HISTORY_HEADERS",
    "BATCH_REQUIRED_HEADERS", "BATCH_OPTIONAL_HEADERS", "BATCH_ALIAS_PAIRS",
    "TEMPLATE_HEADERS", "TEMPLATE_XLSX_SHEETS", "RESULT_XLSX_SHEETS",
    "CONTENT_KINDS", "TICKET_FIELDS", "VerifiedArtifactRef", "StateStore",
    "batch_template_value", "batch_result_value", "semantic_digest_for",
]
